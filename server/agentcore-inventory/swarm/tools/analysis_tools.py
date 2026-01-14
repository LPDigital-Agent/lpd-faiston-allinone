# =============================================================================
# Analysis Tools - File Parsing for Inventory Swarm
# =============================================================================
# Tools for analyzing various file formats (CSV, XLSX, PDF, XML).
#
# Used by: file_analyst agent
# =============================================================================

import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional

from strands import tool

logger = logging.getLogger(__name__)


@tool
def detect_file_type(file_path: str) -> Dict[str, Any]:
    """
    Detect the type of an uploaded file based on extension and content.

    Args:
        file_path: Path to the file (local or S3)

    Returns:
        dict with:
        - file_type: "csv", "xlsx", "pdf", "xml", or "unknown"
        - extension: File extension
        - mime_type: Detected MIME type
        - confidence: Detection confidence (0.0-1.0)
    """
    logger.info("[detect_file_type] Analyzing: %s", file_path)

    # Get extension
    path = Path(file_path)
    extension = path.suffix.lower()

    # Extension to type mapping
    type_map = {
        ".csv": ("csv", "text/csv", 1.0),
        ".xlsx": ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 1.0),
        ".xls": ("xlsx", "application/vnd.ms-excel", 1.0),
        ".pdf": ("pdf", "application/pdf", 1.0),
        ".xml": ("xml", "application/xml", 1.0),
        ".txt": ("csv", "text/plain", 0.7),  # Might be CSV
    }

    if extension in type_map:
        file_type, mime_type, confidence = type_map[extension]
    else:
        file_type, mime_type, confidence = "unknown", "application/octet-stream", 0.0

    return {
        "file_type": file_type,
        "extension": extension,
        "mime_type": mime_type,
        "confidence": confidence,
        "file_name": path.name,
    }


@tool
def analyze_csv(
    file_path: str,
    encoding: str = "auto",
    delimiter: str = "auto",
    sample_rows: int = 10,
) -> Dict[str, Any]:
    """
    Analyze a CSV file structure and extract column information.

    Args:
        file_path: Path to the CSV file (local or S3)
        encoding: File encoding (auto-detect if "auto")
        delimiter: Column delimiter (auto-detect if "auto")
        sample_rows: Number of sample rows to include

    Returns:
        dict with:
        - columns: List of column info (name, type, samples, nulls)
        - row_count: Total number of rows
        - encoding: Detected encoding
        - delimiter: Detected delimiter
        - overall_confidence: Analysis confidence
        - issues: List of any issues found
    """
    import csv
    import io
    from collections import defaultdict

    logger.info("[analyze_csv] Analyzing: %s", file_path)

    # Load file content
    content = _load_file_content(file_path)

    # Detect encoding if auto
    if encoding == "auto":
        encoding = _detect_encoding(content)
        logger.info("[analyze_csv] Detected encoding: %s", encoding)

    # Decode content
    try:
        text = content.decode(encoding)
    except UnicodeDecodeError:
        text = content.decode("utf-8", errors="replace")
        encoding = "utf-8 (fallback)"

    # Detect delimiter if auto
    if delimiter == "auto":
        delimiter = _detect_csv_delimiter(text)
        logger.info("[analyze_csv] Detected delimiter: %s", repr(delimiter))

    # Parse CSV
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)

    if not rows:
        return {
            "columns": [],
            "row_count": 0,
            "encoding": encoding,
            "delimiter": delimiter,
            "overall_confidence": 0.0,
            "issues": ["Empty file"],
        }

    # Extract headers and data
    headers = rows[0]
    data_rows = rows[1:]
    row_count = len(data_rows)

    # Analyze columns
    columns = []
    for i, header in enumerate(headers):
        col_values = [row[i] if i < len(row) else None for row in data_rows]
        col_info = _analyze_column(header, col_values, sample_rows)
        columns.append(col_info)

    # Calculate overall confidence
    confidences = [c["confidence"] for c in columns]
    overall_confidence = sum(confidences) / len(confidences) if confidences else 0.0

    # Detect issues
    issues = []
    for col in columns:
        if col["null_count"] > row_count * 0.5:
            issues.append(f"Column '{col['name']}' has >50% null values")
        if col["confidence"] < 0.7:
            issues.append(f"Column '{col['name']}' has low type confidence")

    return {
        "columns": columns,
        "row_count": row_count,
        "encoding": encoding,
        "delimiter": delimiter,
        "overall_confidence": overall_confidence,
        "issues": issues,
    }


@tool
def analyze_xlsx(
    file_path: str,
    sheet_name: Optional[str] = None,
    sample_rows: int = 10,
) -> Dict[str, Any]:
    """
    Analyze an Excel file structure and extract column information.

    Args:
        file_path: Path to the XLSX file (local or S3)
        sheet_name: Specific sheet to analyze (first sheet if None)
        sample_rows: Number of sample rows to include

    Returns:
        dict with:
        - columns: List of column info
        - row_count: Total rows
        - sheet_name: Sheet analyzed
        - available_sheets: List of all sheets
        - overall_confidence: Analysis confidence
        - issues: List of any issues found
    """
    logger.info("[analyze_xlsx] Analyzing: %s (sheet=%s)", file_path, sheet_name)

    try:
        import openpyxl
        from io import BytesIO

        content = _load_file_content(file_path)
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

        available_sheets = wb.sheetnames

        # Select sheet
        if sheet_name and sheet_name in available_sheets:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # Read data
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return {
                "columns": [],
                "row_count": 0,
                "sheet_name": sheet_name,
                "available_sheets": available_sheets,
                "overall_confidence": 0.0,
                "issues": ["Empty sheet"],
            }

        # Extract headers and data
        headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]
        row_count = len(data_rows)

        # Analyze columns
        columns = []
        for i, header in enumerate(headers):
            col_values = [row[i] if i < len(row) else None for row in data_rows]
            col_info = _analyze_column(header, col_values, sample_rows)
            columns.append(col_info)

        # Calculate confidence
        confidences = [c["confidence"] for c in columns]
        overall_confidence = sum(confidences) / len(confidences) if confidences else 0.0

        # Detect issues
        issues = []
        if len(available_sheets) > 1:
            issues.append(f"Multiple sheets available: {available_sheets}")

        return {
            "columns": columns,
            "row_count": row_count,
            "sheet_name": sheet_name,
            "available_sheets": available_sheets,
            "overall_confidence": overall_confidence,
            "issues": issues,
        }

    except ImportError:
        logger.error("[analyze_xlsx] openpyxl not installed")
        return {
            "columns": [],
            "row_count": 0,
            "overall_confidence": 0.0,
            "issues": ["openpyxl not installed. Use Meta-Tooling to install."],
        }


@tool
def analyze_pdf(file_path: str, sample_rows: int = 10) -> Dict[str, Any]:
    """
    Analyze a PDF file and extract table data using Gemini Vision.

    Args:
        file_path: Path to the PDF file (local or S3)
        sample_rows: Number of sample rows to include

    Returns:
        dict with:
        - columns: Extracted column info
        - row_count: Estimated rows
        - extraction_method: "vision" or "text"
        - overall_confidence: Analysis confidence
        - issues: List of any issues found
    """
    logger.info("[analyze_pdf] Analyzing: %s", file_path)

    # Note: Full implementation would use Gemini Vision API
    # For now, return placeholder indicating Vision needed
    return {
        "columns": [],
        "row_count": 0,
        "extraction_method": "vision_required",
        "overall_confidence": 0.0,
        "issues": ["PDF analysis requires Gemini Vision. Use the LLM to process."],
        "suggestion": "Use Gemini Pro with the PDF content to extract table data.",
    }


@tool
def analyze_xml(file_path: str, sample_rows: int = 10) -> Dict[str, Any]:
    """
    Analyze an XML file and extract record structure.

    Args:
        file_path: Path to the XML file (local or S3)
        sample_rows: Number of sample records to include

    Returns:
        dict with:
        - columns: Fields found in records
        - row_count: Number of records
        - root_element: XML root element
        - record_element: Repeating record element
        - namespaces: XML namespaces detected
        - overall_confidence: Analysis confidence
        - issues: List of any issues found
    """
    import xml.etree.ElementTree as ET

    logger.info("[analyze_xml] Analyzing: %s", file_path)

    content = _load_file_content(file_path)

    try:
        root = ET.fromstring(content)

        # Detect namespaces
        namespaces = {}
        for elem in root.iter():
            if "}" in elem.tag:
                ns = elem.tag.split("}")[0] + "}"
                prefix = f"ns{len(namespaces)}"
                if ns not in namespaces.values():
                    namespaces[prefix] = ns

        # Find repeating elements (likely records)
        child_counts = {}
        for child in root:
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
            child_counts[tag] = child_counts.get(tag, 0) + 1

        # Most frequent child is likely the record element
        record_element = max(child_counts, key=child_counts.get) if child_counts else None
        row_count = child_counts.get(record_element, 0) if record_element else 0

        # Extract columns from first record
        columns = []
        if record_element:
            for record in root.findall(f".//{record_element}"):
                for field in record:
                    tag = field.tag.split("}")[-1] if "}" in field.tag else field.tag
                    if not any(c["name"] == tag for c in columns):
                        columns.append({
                            "name": tag,
                            "inferred_type": _infer_type(field.text),
                            "sample_values": [field.text] if field.text else [],
                            "null_count": 0 if field.text else 1,
                            "confidence": 0.8,
                        })
                break

        return {
            "columns": columns,
            "row_count": row_count,
            "root_element": root.tag.split("}")[-1] if "}" in root.tag else root.tag,
            "record_element": record_element,
            "namespaces": namespaces,
            "overall_confidence": 0.85 if columns else 0.0,
            "issues": [],
        }

    except ET.ParseError as e:
        return {
            "columns": [],
            "row_count": 0,
            "overall_confidence": 0.0,
            "issues": [f"XML parse error: {e}"],
        }


# =============================================================================
# Helper Functions
# =============================================================================


def _load_file_content(file_path: str) -> bytes:
    """Load file content from local path or S3."""
    if file_path.startswith("s3://"):
        import boto3

        s3 = boto3.client("s3")
        bucket, key = file_path[5:].split("/", 1)
        response = s3.get_object(Bucket=bucket, Key=key)
        return response["Body"].read()
    else:
        with open(file_path, "rb") as f:
            return f.read()


def _detect_encoding(content: bytes) -> str:
    """Detect file encoding."""
    # Simple BOM detection
    if content.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if content.startswith(b"\xff\xfe"):
        return "utf-16-le"
    if content.startswith(b"\xfe\xff"):
        return "utf-16-be"

    # Try chardet if available
    try:
        import chardet

        result = chardet.detect(content[:10000])
        return result.get("encoding", "utf-8") or "utf-8"
    except ImportError:
        return "utf-8"


def _detect_csv_delimiter(text: str) -> str:
    """Detect CSV delimiter from content."""
    import csv

    try:
        dialect = csv.Sniffer().sniff(text[:5000])
        return dialect.delimiter
    except csv.Error:
        # Count common delimiters
        delimiters = [",", ";", "\t", "|"]
        counts = {d: text.count(d) for d in delimiters}
        return max(counts, key=counts.get)


def _analyze_column(
    name: str,
    values: List[Any],
    sample_rows: int,
) -> Dict[str, Any]:
    """Analyze a single column's type and statistics."""
    # Filter non-null values
    non_null = [v for v in values if v is not None and str(v).strip()]
    null_count = len(values) - len(non_null)

    # Get sample values
    samples = [str(v) for v in non_null[:sample_rows]]

    # Infer type
    inferred_type = _infer_column_type(non_null)

    # Calculate confidence based on type consistency
    if non_null:
        type_matches = sum(1 for v in non_null if _matches_type(v, inferred_type))
        confidence = type_matches / len(non_null)
    else:
        confidence = 0.0

    return {
        "name": name,
        "inferred_type": inferred_type,
        "sample_values": samples,
        "null_count": null_count,
        "unique_count": len(set(str(v) for v in non_null)),
        "confidence": round(confidence, 2),
    }


def _infer_column_type(values: List[Any]) -> str:
    """Infer the data type of a column."""
    if not values:
        return "unknown"

    # Sample for type detection
    sample = values[:100]

    # Count type matches
    type_counts = {"integer": 0, "decimal": 0, "date": 0, "string": 0}

    for v in sample:
        str_val = str(v).strip()
        if _is_integer(str_val):
            type_counts["integer"] += 1
        elif _is_decimal(str_val):
            type_counts["decimal"] += 1
        elif _is_date(str_val):
            type_counts["date"] += 1
        else:
            type_counts["string"] += 1

    return max(type_counts, key=type_counts.get)


def _infer_type(value: Optional[str]) -> str:
    """Infer type of a single value."""
    if value is None:
        return "unknown"
    if _is_integer(value):
        return "integer"
    if _is_decimal(value):
        return "decimal"
    if _is_date(value):
        return "date"
    return "string"


def _is_integer(val: str) -> bool:
    """Check if value is integer."""
    try:
        int(val.replace(",", "").replace(".", ""))
        return "." not in val and "," not in val
    except (ValueError, AttributeError):
        return False


def _is_decimal(val: str) -> bool:
    """Check if value is decimal."""
    try:
        float(val.replace(",", "."))
        return True
    except (ValueError, AttributeError):
        return False


def _is_date(val: str) -> bool:
    """Check if value looks like a date."""
    import re

    date_patterns = [
        r"\d{4}-\d{2}-\d{2}",  # ISO
        r"\d{2}/\d{2}/\d{4}",  # DD/MM/YYYY or MM/DD/YYYY
        r"\d{2}-\d{2}-\d{4}",  # DD-MM-YYYY
    ]
    return any(re.match(p, str(val)) for p in date_patterns)


def _matches_type(value: Any, expected_type: str) -> bool:
    """Check if value matches expected type."""
    str_val = str(value).strip()
    if expected_type == "integer":
        return _is_integer(str_val)
    if expected_type == "decimal":
        return _is_decimal(str_val)
    if expected_type == "date":
        return _is_date(str_val)
    return True  # String matches everything
