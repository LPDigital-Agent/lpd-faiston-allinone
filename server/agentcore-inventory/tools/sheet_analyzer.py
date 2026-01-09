# =============================================================================
# Sheet Analyzer Tool - Multi-Sheet XLSX Analysis
# =============================================================================
# Analyzes XLSX files with multiple sheets to understand structure,
# detect relationships, and suggest processing strategies.
#
# Philosophy: OBSERVE → THINK → ASK → LEARN → ACT
# This tool handles the OBSERVE phase by deeply analyzing file structure.
#
# SCHEMA-AWARE: Now uses SchemaColumnMatcher for dynamic column matching
# against PostgreSQL schema instead of hardcoded patterns.
#
# Module: Gestao de Ativos -> Gestao de Estoque -> Smart Import
# Author: Faiston NEXO Team
# Updated: January 2026 - Schema-aware matching
# =============================================================================

import io
import logging
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum

logger = logging.getLogger(__name__)


# =============================================================================
# Types and Enums
# =============================================================================


class SheetPurpose(Enum):
    """Detected purpose of a sheet based on content analysis."""
    ITEMS = "items"                 # Main items/products list
    SERIALS = "serials"             # Serial numbers list
    METADATA = "metadata"           # File metadata/header info
    SUMMARY = "summary"             # Summary/totals sheet
    PROJECTS = "projects"           # Project assignments
    LOCATIONS = "locations"         # Location/warehouse data
    UNKNOWN = "unknown"             # Purpose not detected


class SheetRelationship(Enum):
    """Relationship type between sheets."""
    MASTER_DETAIL = "master_detail"     # Sheet1 is header, Sheet2 is details
    ONE_TO_MANY = "one_to_many"         # Sheet1 items have multiple Sheet2 records
    SEPARATE = "separate"               # Independent sheets
    DUPLICATE = "duplicate"             # Same data, different format
    COMPLEMENT = "complement"           # Sheets complement each other


@dataclass
class ColumnAnalysis:
    """Analysis of a single column."""
    name: str
    normalized_name: str
    sample_values: List[str]
    data_type: str              # "text", "number", "date", "mixed"
    unique_count: int
    null_count: int
    is_likely_key: bool         # Appears to be an ID/key column
    suggested_mapping: Optional[str] = None
    mapping_confidence: float = 0.0


@dataclass
class SheetAnalysis:
    """Analysis of a single sheet."""
    name: str
    row_count: int
    column_count: int
    columns: List[ColumnAnalysis]
    detected_purpose: SheetPurpose
    purpose_confidence: float
    has_headers: bool
    suggested_action: str       # "process", "skip", "merge_with"
    merge_target: Optional[str] = None  # If merge, which sheet
    notes: List[str] = field(default_factory=list)


@dataclass
class SheetRelationshipAnalysis:
    """Analysis of relationship between two sheets."""
    sheet1: str
    sheet2: str
    relationship_type: SheetRelationship
    confidence: float
    join_columns: List[Tuple[str, str]]  # (sheet1_col, sheet2_col)
    description: str


@dataclass
class WorkbookAnalysis:
    """Complete analysis of an XLSX workbook."""
    filename: str
    sheet_count: int
    total_rows: int
    sheets: List[SheetAnalysis]
    relationships: List[SheetRelationshipAnalysis]
    recommended_strategy: str
    questions_for_user: List[Dict[str, Any]]
    reasoning_trace: List[Dict[str, str]]


# =============================================================================
# Column Pattern Detection (OBSERVE phase)
# Schema-aware matching using SchemaColumnMatcher
# =============================================================================


# Global matcher instance (lazy initialized)
_schema_matcher = None


def _get_schema_matcher():
    """
    Get or create the SchemaColumnMatcher instance (lazy initialization).

    Uses lazy import to avoid cold start impact in AgentCore runtime.
    Falls back to pattern-only matching if schema is unavailable.
    """
    global _schema_matcher
    if _schema_matcher is None:
        try:
            from tools.schema_column_matcher import get_column_matcher
            _schema_matcher = get_column_matcher()
            logger.info("[SheetAnalyzer] Schema-aware matching enabled")
        except Exception as e:
            logger.warning(f"[SheetAnalyzer] Schema matcher unavailable: {e}")
            _schema_matcher = None
    return _schema_matcher


# Legacy COLUMN_PATTERNS kept as fallback when schema is unavailable
# These patterns are now also in schema_column_matcher.py BUILTIN_ALIASES
COLUMN_PATTERNS: Dict[str, List[str]] = {
    "part_number": [
        "pn", "codigo", "material", "part_number", "partnumber",
        "cod_material", "codigo_material", "item", "sku", "codigo_item",
        "equipamento", "equipment", "cod_equip"
    ],
    "description": [
        "desc", "descricao", "nome", "description", "nome_material",
        "desc_material", "produto", "product", "item_name"
    ],
    "quantity": [
        "qty", "quantidade", "qtd", "quant", "quantity", "qtde"
    ],
    "serial_number": [
        "serial", "sn", "serie", "serial_number", "numero_serie",
        "ns", "n_serie"
    ],
    "location_code": [
        "loc", "local", "deposito", "location", "warehouse",
        "armazem", "almoxarifado", "destino"
    ],
    "project_code": [
        "projeto", "project", "proj", "id_projeto", "cod_projeto",
        "project_id", "obra"
    ],
    "supplier_name": [
        "fornecedor", "supplier", "vendor", "fabricante", "manufacturer"
    ],
    "unit_value": [
        "custo", "cost", "preco", "valor", "unit_cost", "custo_unitario",
        "preco_unitario", "value"
    ],
    "ncm": [
        "ncm", "ncm_code", "codigo_ncm"
    ],
    "movement_date": [
        "data", "date", "dt", "data_entrada", "data_saida", "created_at"
    ],
    "status": [
        "status", "situacao", "estado", "state"
    ],
}


def normalize_column_name(name: str) -> str:
    """
    Normalize column name for pattern matching.

    Args:
        name: Original column name

    Returns:
        Normalized lowercase name without accents/special chars
    """
    if not name:
        return ""

    # Accent replacements
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "è": "e", "ê": "e",
        "í": "i", "ì": "i", "î": "i",
        "ó": "o", "ò": "o", "õ": "o", "ô": "o",
        "ú": "u", "ù": "u", "û": "u",
        "ç": "c", "ñ": "n",
    }

    result = str(name).lower().strip()
    for char, replacement in replacements.items():
        result = result.replace(char, replacement)

    # Keep only alphanumeric and underscore
    result = "".join(c if c.isalnum() or c == "_" else "_" for c in result)

    # Remove multiple underscores
    while "__" in result:
        result = result.replace("__", "_")

    return result.strip("_")


def detect_column_mapping(
    column_name: str,
    target_table: str = "pending_entry_items",
) -> Tuple[Optional[str], float]:
    """
    Detect which target field a column maps to using schema-aware matching.

    Algorithm (priority order):
    1. Use SchemaColumnMatcher (schema + aliases + fuzzy) if available
    2. Fallback to legacy COLUMN_PATTERNS if schema unavailable

    Args:
        column_name: Original column name
        target_table: Target PostgreSQL table (default: pending_entry_items)

    Returns:
        Tuple of (target_field or None, confidence)
    """
    # Try schema-aware matching first
    matcher = _get_schema_matcher()
    if matcher:
        target, confidence = matcher.match_column(column_name, target_table)
        if target:
            logger.debug(
                f"[SheetAnalyzer] Schema match: {column_name} → {target} ({confidence:.2f})"
            )
            return target, confidence

    # Fallback to legacy pattern matching
    normalized = normalize_column_name(column_name)

    for target_field, patterns in COLUMN_PATTERNS.items():
        # Exact match
        normalized_patterns = [normalize_column_name(p) for p in patterns]
        if normalized in normalized_patterns:
            return target_field, 0.95

        # Contains match
        for pattern in normalized_patterns:
            if pattern in normalized or normalized in pattern:
                return target_field, 0.75

    return None, 0.0


def detect_column_mapping_with_schema(
    column_name: str,
    target_table: str,
    schema_provider=None,
) -> Tuple[Optional[str], float]:
    """
    Schema-aware column mapping detection (explicit schema injection).

    Use this when you have a SchemaProvider instance and want to ensure
    schema awareness without relying on global state.

    Args:
        column_name: Original column name
        target_table: Target PostgreSQL table
        schema_provider: SchemaProvider instance (optional)

    Returns:
        Tuple of (target_field or None, confidence)
    """
    try:
        from tools.schema_column_matcher import SchemaColumnMatcher
        matcher = SchemaColumnMatcher(schema_provider)
        return matcher.match_column(column_name, target_table)
    except Exception as e:
        logger.warning(f"[SheetAnalyzer] Schema match failed: {e}")
        return detect_column_mapping(column_name, target_table)


def detect_data_type(values: List[Any]) -> str:
    """
    Detect the predominant data type in a list of values.

    Args:
        values: Sample values from column

    Returns:
        "text", "number", "date", or "mixed"
    """
    if not values:
        return "text"

    type_counts = {"text": 0, "number": 0, "date": 0, "empty": 0}

    for val in values:
        if val is None or str(val).strip() == "":
            type_counts["empty"] += 1
            continue

        str_val = str(val).strip()

        # Check if number
        try:
            float(str_val.replace(",", ".").replace(" ", ""))
            type_counts["number"] += 1
            continue
        except ValueError:
            pass

        # Check if date-like (simple patterns)
        if any(sep in str_val for sep in ["/", "-"]) and len(str_val) <= 20:
            if any(c.isdigit() for c in str_val):
                type_counts["date"] += 1
                continue

        type_counts["text"] += 1

    # Determine predominant type
    non_empty = sum(v for k, v in type_counts.items() if k != "empty")
    if non_empty == 0:
        return "text"

    for dtype, count in type_counts.items():
        if dtype != "empty" and count / non_empty > 0.7:
            return dtype

    return "mixed"


# =============================================================================
# Sheet Purpose Detection (THINK phase)
# =============================================================================


def detect_sheet_purpose(
    sheet_name: str,
    columns: List[ColumnAnalysis],
    row_count: int,
) -> Tuple[SheetPurpose, float]:
    """
    Detect the purpose of a sheet based on name and content.

    Args:
        sheet_name: Name of the sheet
        columns: Analyzed columns
        row_count: Number of data rows

    Returns:
        Tuple of (SheetPurpose, confidence)
    """
    name_lower = normalize_column_name(sheet_name)

    # Check sheet name patterns
    name_hints = {
        SheetPurpose.SERIALS: ["serial", "serie", "sn", "ns"],
        SheetPurpose.SUMMARY: ["resumo", "summary", "total", "consolidado"],
        SheetPurpose.METADATA: ["info", "cabecalho", "header", "meta"],
        SheetPurpose.PROJECTS: ["projeto", "project", "obra"],
        SheetPurpose.LOCATIONS: ["local", "location", "deposito", "armazem"],
        SheetPurpose.ITEMS: ["item", "material", "produto", "equipamento", "dados"],
    }

    for purpose, hints in name_hints.items():
        if any(h in name_lower for h in hints):
            return purpose, 0.85

    # Analyze columns to determine purpose
    column_fields = [c.suggested_mapping for c in columns if c.suggested_mapping]

    # Items sheet usually has part_number, quantity, description
    items_indicators = {"part_number", "quantity", "description"}
    if len(items_indicators & set(column_fields)) >= 2:
        return SheetPurpose.ITEMS, 0.90

    # Serials sheet has mostly serial columns
    has_serial = "serial" in column_fields
    few_columns = len(columns) <= 5
    if has_serial and few_columns:
        return SheetPurpose.SERIALS, 0.80

    # Summary sheet usually has few rows and aggregated data
    if row_count <= 10 and any("total" in normalize_column_name(c.name) for c in columns):
        return SheetPurpose.SUMMARY, 0.75

    # Default to ITEMS for sheets with many rows
    if row_count > 10:
        return SheetPurpose.ITEMS, 0.50

    return SheetPurpose.UNKNOWN, 0.30


# =============================================================================
# Relationship Detection (THINK phase)
# =============================================================================


def detect_sheet_relationships(
    sheets: List[SheetAnalysis],
) -> List[SheetRelationshipAnalysis]:
    """
    Detect relationships between sheets.

    Args:
        sheets: List of analyzed sheets

    Returns:
        List of relationship analyses
    """
    relationships = []

    for i, sheet1 in enumerate(sheets):
        for sheet2 in sheets[i + 1:]:
            relationship = _analyze_sheet_pair(sheet1, sheet2)
            if relationship:
                relationships.append(relationship)

    return relationships


def _analyze_sheet_pair(
    sheet1: SheetAnalysis,
    sheet2: SheetAnalysis,
) -> Optional[SheetRelationshipAnalysis]:
    """Analyze relationship between two sheets."""
    # Find common column mappings
    sheet1_mappings = {c.suggested_mapping for c in sheet1.columns if c.suggested_mapping}
    sheet2_mappings = {c.suggested_mapping for c in sheet2.columns if c.suggested_mapping}
    common_mappings = sheet1_mappings & sheet2_mappings

    if not common_mappings:
        return SheetRelationshipAnalysis(
            sheet1=sheet1.name,
            sheet2=sheet2.name,
            relationship_type=SheetRelationship.SEPARATE,
            confidence=0.70,
            join_columns=[],
            description=f"Abas '{sheet1.name}' e '{sheet2.name}' parecem independentes",
        )

    # Check for master-detail pattern
    if sheet1.detected_purpose == SheetPurpose.ITEMS:
        if sheet2.detected_purpose == SheetPurpose.SERIALS:
            join_cols = _find_join_columns(sheet1, sheet2, common_mappings)
            return SheetRelationshipAnalysis(
                sheet1=sheet1.name,
                sheet2=sheet2.name,
                relationship_type=SheetRelationship.ONE_TO_MANY,
                confidence=0.85,
                join_columns=join_cols,
                description=f"'{sheet1.name}' contém itens, '{sheet2.name}' contém seriais relacionados",
            )

    # Check for complement pattern
    if sheet1.detected_purpose == sheet2.detected_purpose:
        return SheetRelationshipAnalysis(
            sheet1=sheet1.name,
            sheet2=sheet2.name,
            relationship_type=SheetRelationship.COMPLEMENT,
            confidence=0.60,
            join_columns=_find_join_columns(sheet1, sheet2, common_mappings),
            description=f"Abas '{sheet1.name}' e '{sheet2.name}' podem complementar dados",
        )

    return None


def _find_join_columns(
    sheet1: SheetAnalysis,
    sheet2: SheetAnalysis,
    common_mappings: set,
) -> List[Tuple[str, str]]:
    """Find columns that can be used to join two sheets."""
    join_cols = []

    for mapping in common_mappings:
        col1 = next((c for c in sheet1.columns if c.suggested_mapping == mapping), None)
        col2 = next((c for c in sheet2.columns if c.suggested_mapping == mapping), None)

        if col1 and col2:
            join_cols.append((col1.name, col2.name))

    return join_cols


# =============================================================================
# Question Generation (ASK phase)
# =============================================================================


def generate_questions(
    analysis: "WorkbookAnalysis",
) -> List[Dict[str, Any]]:
    """
    Generate clarification questions for the user based on analysis.

    Args:
        analysis: Complete workbook analysis

    Returns:
        List of question dictionaries
    """
    questions = []

    # Question about multiple sheets
    if analysis.sheet_count > 1:
        sheet_options = [
            {"value": "all", "label": "Processar todas as abas"},
            {"value": "first", "label": f"Apenas a primeira aba ({analysis.sheets[0].name})"},
            {"value": "select", "label": "Deixe-me escolher quais abas"},
        ]

        questions.append({
            "id": "sheet_selection",
            "question": f"Encontrei {analysis.sheet_count} abas neste arquivo. Como devo processar?",
            "context": _format_sheets_context(analysis.sheets),
            "options": sheet_options,
            "importance": "high",
            "topic": "sheet_selection",
        })

    # Question about unmapped columns
    for sheet in analysis.sheets:
        unmapped = [c for c in sheet.columns if not c.suggested_mapping]
        if len(unmapped) > 5:
            # Find columns that might be important
            potential_keys = [c for c in unmapped if c.is_likely_key]
            if potential_keys:
                questions.append({
                    "id": f"column_mapping_{sheet.name}",
                    "question": f"Na aba '{sheet.name}', a coluna '{potential_keys[0].name}' corresponde a qual campo?",
                    "context": f"Valores exemplo: {', '.join(potential_keys[0].sample_values[:3])}",
                    "options": [
                        {"value": "part_number", "label": "Part Number / Código"},
                        {"value": "serial", "label": "Número de Série"},
                        {"value": "project", "label": "Projeto"},
                        {"value": "ignore", "label": "Ignorar esta coluna"},
                    ],
                    "importance": "critical",
                    "topic": "column_mapping",
                    "column": potential_keys[0].name,
                })

    # Question about movement type (if not obvious)
    questions.append({
        "id": "movement_type",
        "question": "Este arquivo é para qual tipo de movimentação?",
        "context": "Preciso saber o tipo para processar corretamente",
        "options": [
            {"value": "entry", "label": "Entrada (Internalização)"},
            {"value": "exit", "label": "Saída (Expedição)"},
            {"value": "transfer", "label": "Transferência"},
            {"value": "reverse", "label": "Reversa (Devolução)"},
        ],
        "importance": "critical",
        "topic": "movement_type",
    })

    return questions


def _format_sheets_context(sheets: List[SheetAnalysis]) -> str:
    """Format sheets info for question context."""
    lines = []
    for sheet in sheets:
        purpose_label = {
            SheetPurpose.ITEMS: "Itens",
            SheetPurpose.SERIALS: "Seriais",
            SheetPurpose.SUMMARY: "Resumo",
            SheetPurpose.METADATA: "Metadados",
            SheetPurpose.PROJECTS: "Projetos",
            SheetPurpose.LOCATIONS: "Locais",
            SheetPurpose.UNKNOWN: "Desconhecido",
        }.get(sheet.detected_purpose, "?")

        lines.append(f"• {sheet.name}: {sheet.row_count} linhas ({purpose_label})")

    return "\n".join(lines)


# =============================================================================
# Main Analysis Function (OBSERVE + THINK)
# =============================================================================


def analyze_workbook(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Perform complete analysis of an XLSX workbook.

    This is the main entry point for the OBSERVE and THINK phases.

    Args:
        content: Raw file content as bytes
        filename: Original filename
        max_sample_rows: Maximum rows to sample per sheet

    Returns:
        Complete WorkbookAnalysis with recommendations
    """
    # Import openpyxl (lazy import to reduce cold start)
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl package required for Excel analysis. "
            "Add 'openpyxl>=3.1.0' to requirements.txt"
        )

    reasoning_trace = []

    # OBSERVE: Load workbook
    reasoning_trace.append({
        "type": "thought",
        "content": f"Vou analisar a estrutura do arquivo '{filename}'",
    })

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    reasoning_trace.append({
        "type": "observation",
        "content": f"Arquivo tem {len(wb.sheetnames)} aba(s): {', '.join(wb.sheetnames)}",
    })

    # OBSERVE: Analyze each sheet
    sheets_analysis = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        reasoning_trace.append({
            "type": "action",
            "content": f"Analisando aba '{sheet_name}'",
        })

        ws = wb[sheet_name]
        sheet_analysis = _analyze_sheet(ws, sheet_name, max_sample_rows)
        sheets_analysis.append(sheet_analysis)
        total_rows += sheet_analysis.row_count

        reasoning_trace.append({
            "type": "observation",
            "content": (
                f"Aba '{sheet_name}': {sheet_analysis.row_count} linhas, "
                f"{sheet_analysis.column_count} colunas, "
                f"propósito detectado: {sheet_analysis.detected_purpose.value}"
            ),
        })

    wb.close()

    # THINK: Detect relationships
    reasoning_trace.append({
        "type": "thought",
        "content": "Vou verificar se existe relação entre as abas",
    })

    relationships = detect_sheet_relationships(sheets_analysis)

    for rel in relationships:
        reasoning_trace.append({
            "type": "observation",
            "content": rel.description,
        })

    # THINK: Determine strategy
    reasoning_trace.append({
        "type": "thought",
        "content": "Decidindo a melhor estratégia de processamento",
    })

    strategy = _determine_processing_strategy(sheets_analysis, relationships)

    reasoning_trace.append({
        "type": "conclusion",
        "content": f"Estratégia recomendada: {strategy}",
    })

    # Build analysis
    analysis = WorkbookAnalysis(
        filename=filename,
        sheet_count=len(wb.sheetnames),
        total_rows=total_rows,
        sheets=sheets_analysis,
        relationships=relationships,
        recommended_strategy=strategy,
        questions_for_user=[],
        reasoning_trace=reasoning_trace,
    )

    # ASK: Generate questions
    analysis.questions_for_user = generate_questions(analysis)

    return analysis


def _analyze_sheet(
    ws,
    sheet_name: str,
    max_sample_rows: int,
) -> SheetAnalysis:
    """Analyze a single worksheet."""
    rows_iter = ws.iter_rows(values_only=True)

    # Get headers
    headers_row = next(rows_iter, None)
    if not headers_row:
        return SheetAnalysis(
            name=sheet_name,
            row_count=0,
            column_count=0,
            columns=[],
            detected_purpose=SheetPurpose.UNKNOWN,
            purpose_confidence=0.0,
            has_headers=False,
            suggested_action="skip",
            notes=["Aba vazia"],
        )

    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers_row)]

    # Sample rows for analysis
    sample_rows = []
    row_count = 0
    for row in rows_iter:
        row_count += 1
        if len(sample_rows) < max_sample_rows:
            sample_rows.append(row)

    # Analyze columns
    columns = []
    for col_idx, header in enumerate(headers):
        col_values = [row[col_idx] if col_idx < len(row) else None for row in sample_rows]

        # Detect mapping
        mapping, confidence = detect_column_mapping(header)

        # Detect data type
        data_type = detect_data_type(col_values)

        # Detect if likely key column
        unique_vals = set(str(v) for v in col_values if v)
        is_key = len(unique_vals) == len([v for v in col_values if v])

        columns.append(ColumnAnalysis(
            name=header,
            normalized_name=normalize_column_name(header),
            sample_values=[str(v) for v in col_values[:5] if v],
            data_type=data_type,
            unique_count=len(unique_vals),
            null_count=sum(1 for v in col_values if not v),
            is_likely_key=is_key and len(unique_vals) > 1,
            suggested_mapping=mapping,
            mapping_confidence=confidence,
        ))

    # Detect sheet purpose
    purpose, purpose_confidence = detect_sheet_purpose(sheet_name, columns, row_count)

    # Determine action
    if purpose == SheetPurpose.SUMMARY or purpose == SheetPurpose.METADATA:
        action = "skip"
    elif purpose == SheetPurpose.SERIALS:
        action = "merge_with"
    else:
        action = "process"

    return SheetAnalysis(
        name=sheet_name,
        row_count=row_count,
        column_count=len(headers),
        columns=columns,
        detected_purpose=purpose,
        purpose_confidence=purpose_confidence,
        has_headers=True,
        suggested_action=action,
        notes=[],
    )


def _determine_processing_strategy(
    sheets: List[SheetAnalysis],
    relationships: List[SheetRelationshipAnalysis],
) -> str:
    """Determine the best processing strategy for the workbook."""
    # Single sheet - straightforward
    if len(sheets) == 1:
        return "process_single"

    # All sheets are items - process all
    items_sheets = [s for s in sheets if s.detected_purpose == SheetPurpose.ITEMS]
    if len(items_sheets) == len(sheets):
        return "process_all_separate"

    # Has items + serials relationship
    has_serial_rel = any(
        r.relationship_type == SheetRelationship.ONE_TO_MANY
        for r in relationships
    )
    if has_serial_rel:
        return "merge_items_serials"

    # Has complement relationship
    has_complement = any(
        r.relationship_type == SheetRelationship.COMPLEMENT
        for r in relationships
    )
    if has_complement:
        return "merge_complement"

    # Default: process main sheet only
    return "process_main_only"


# =============================================================================
# Serialization for JSON Response
# =============================================================================


def analysis_to_dict(analysis: WorkbookAnalysis) -> Dict[str, Any]:
    """
    Convert WorkbookAnalysis to dictionary for JSON serialization.

    Args:
        analysis: WorkbookAnalysis object

    Returns:
        Dictionary representation
    """
    return {
        "filename": analysis.filename,
        "sheet_count": analysis.sheet_count,
        "total_rows": analysis.total_rows,
        "sheets": [
            {
                "name": s.name,
                "row_count": s.row_count,
                "column_count": s.column_count,
                "columns": [
                    {
                        "name": c.name,
                        "normalized_name": c.normalized_name,
                        "sample_values": c.sample_values,
                        "data_type": c.data_type,
                        "unique_count": c.unique_count,
                        "null_count": c.null_count,
                        "is_likely_key": c.is_likely_key,
                        "suggested_mapping": c.suggested_mapping,
                        "mapping_confidence": c.mapping_confidence,
                    }
                    for c in s.columns
                ],
                "purpose": s.detected_purpose.value,
                "confidence": s.purpose_confidence,
                "has_headers": s.has_headers,
                "suggested_action": s.suggested_action,
                "merge_target": s.merge_target,
                "notes": s.notes,
            }
            for s in analysis.sheets
        ],
        "relationships": [
            {
                "sheet1": r.sheet1,
                "sheet2": r.sheet2,
                "relationship_type": r.relationship_type.value,
                "confidence": r.confidence,
                "join_columns": r.join_columns,
                "description": r.description,
            }
            for r in analysis.relationships
        ],
        "recommended_strategy": analysis.recommended_strategy,
        "questions_for_user": analysis.questions_for_user,
        "reasoning_trace": analysis.reasoning_trace,
    }


# =============================================================================
# Row Aggregation by Part Number (January 2026 Feature)
# =============================================================================
# When a CSV/XLSX file doesn't have a quantity column, we can aggregate rows
# by part_number and use the count as quantity. This is common for equipment
# lists where each row represents a single unit.


@dataclass
class AggregationResult:
    """Result of row aggregation by part number."""

    original_row_count: int
    aggregated_row_count: int
    aggregated_rows: List[Dict[str, Any]]
    part_number_column: str
    aggregation_applied: bool
    duplicate_part_numbers: List[str]


def detect_aggregation_need(
    analysis: WorkbookAnalysis,
    parsed_rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Detect if aggregation by part_number is needed.

    Criteria for aggregation:
    1. No 'quantity' column detected in the analysis
    2. 'part_number' column exists
    3. There are duplicate part numbers in the data

    Args:
        analysis: WorkbookAnalysis from sheet analysis
        parsed_rows: Parsed row data (list of dicts)

    Returns:
        Detection result with recommendation:
        {
            "needs_aggregation": bool,
            "reason": str,
            "part_number_column": str | None,
            "unique_parts": int,
            "total_rows": int,
            "duplicates_sample": list[str],
        }
    """
    # Find quantity and part_number mappings from analysis
    has_quantity = False
    pn_column = None

    for sheet in analysis.sheets:
        for col in sheet.columns:
            if col.suggested_mapping == "quantity":
                has_quantity = True
            elif col.suggested_mapping == "part_number":
                pn_column = col.name

    # If quantity exists or no part_number column, no aggregation needed
    if has_quantity:
        return {
            "needs_aggregation": False,
            "reason": "quantity column exists",
            "part_number_column": pn_column,
        }

    if not pn_column:
        return {
            "needs_aggregation": False,
            "reason": "no part_number column detected",
            "part_number_column": None,
        }

    # Count part numbers to detect duplicates
    pn_counts: Dict[str, int] = {}
    for row in parsed_rows:
        # Try both the original column name and normalized "part_number"
        pn = row.get(pn_column) or row.get("part_number")
        if pn:
            pn_str = str(pn).strip()
            if pn_str:
                pn_counts[pn_str] = pn_counts.get(pn_str, 0) + 1

    # Find duplicates (count > 1)
    duplicates = [pn for pn, count in pn_counts.items() if count > 1]

    return {
        "needs_aggregation": len(duplicates) > 0,
        "reason": (
            f"found {len(duplicates)} duplicate part numbers without quantity column"
            if duplicates
            else "no duplicate part numbers"
        ),
        "part_number_column": pn_column,
        "unique_parts": len(pn_counts),
        "total_rows": len(parsed_rows),
        "duplicates_sample": duplicates[:5],  # First 5 duplicates as sample
    }


def aggregate_rows_by_part_number(
    rows: List[Dict[str, Any]],
    pn_column: str,
    merge_strategy: str = "first",
) -> AggregationResult:
    """
    Aggregate rows by part number, counting occurrences as quantity.

    Each unique part_number becomes one row with quantity = count of occurrences.
    Other columns are merged according to the merge_strategy.

    Args:
        rows: List of parsed row dictionaries
        pn_column: Column name containing part number
        merge_strategy: How to merge other columns:
            - "first": Use first occurrence's values (default)
            - "last": Use last occurrence's values
            - "most_common": Use most common value for each column

    Returns:
        AggregationResult with aggregated rows
    """
    from collections import defaultdict, Counter

    # Group rows by part number
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        # Try both the original column name and normalized "part_number"
        pn = str(row.get(pn_column) or row.get("part_number") or "").strip()
        if pn:
            grouped[pn].append(row)

    aggregated: List[Dict[str, Any]] = []
    duplicates: List[str] = []

    for pn, group_rows in grouped.items():
        if len(group_rows) > 1:
            duplicates.append(pn)

        # Merge row data based on strategy
        merged: Dict[str, Any] = {}

        if merge_strategy == "first":
            merged = dict(group_rows[0])
        elif merge_strategy == "last":
            merged = dict(group_rows[-1])
        elif merge_strategy == "most_common":
            # For each column, use the most common non-empty value
            all_keys: set = set()
            for r in group_rows:
                all_keys.update(r.keys())

            for key in all_keys:
                values = [r.get(key) for r in group_rows if r.get(key)]
                if values:
                    counter = Counter(values)
                    merged[key] = counter.most_common(1)[0][0]
        else:
            # Default to first
            merged = dict(group_rows[0])

        # Set quantity as count of rows (the key feature)
        merged["quantity"] = len(group_rows)
        merged["_aggregated_from_rows"] = len(group_rows)  # Metadata for debugging
        merged["part_number"] = pn  # Ensure part_number is always set

        aggregated.append(merged)

    return AggregationResult(
        original_row_count=len(rows),
        aggregated_row_count=len(aggregated),
        aggregated_rows=aggregated,
        part_number_column=pn_column,
        aggregation_applied=True,
        duplicate_part_numbers=duplicates,
    )


def validate_unique_part_numbers(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Validate that all part numbers are unique after aggregation.

    Args:
        rows: Aggregated rows to validate

    Returns:
        Validation result:
        {
            "is_valid": bool,
            "unique_count": int,
            "duplicate_errors": list of {part_number, first_row, duplicate_row}
        }
    """
    pn_seen: Dict[str, int] = {}
    duplicates: List[Dict[str, Any]] = []

    for i, row in enumerate(rows):
        pn = str(row.get("part_number", "")).strip()
        if pn in pn_seen:
            duplicates.append({
                "part_number": pn,
                "first_row": pn_seen[pn],
                "duplicate_row": i,
            })
        else:
            pn_seen[pn] = i

    return {
        "is_valid": len(duplicates) == 0,
        "unique_count": len(pn_seen),
        "duplicate_errors": duplicates,
    }


# =============================================================================
# Smart File Analysis (FIX January 2026)
# =============================================================================
# Routes to the correct parser based on file type detection.
# This is the AI-First approach: detect file type intelligently and activate
# the appropriate handler.


def analyze_file_smart(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Intelligently analyze any supported file format.

    Supported formats:
    - Structured: CSV, XLSX, XLS, TXT, JSON
    - Visual (AI-First via Gemini Vision): JPG, JPEG, PNG, PDF

    This is the AI-First entry point that:
    1. Detects file type from extension AND content (magic bytes)
    2. Routes to the appropriate parser
    3. Returns a unified WorkbookAnalysis structure

    Args:
        content: Raw file content as bytes
        filename: Original filename (used for type detection)
        max_sample_rows: Maximum rows to sample per sheet

    Returns:
        WorkbookAnalysis with unified structure regardless of file type
    """
    lower_name = filename.lower()

    # ==========================================================================
    # STRUCTURED FORMATS (Direct Parsing)
    # ==========================================================================
    if lower_name.endswith(".csv"):
        return _analyze_csv(content, filename, max_sample_rows)

    elif lower_name.endswith(".xlsx"):
        return analyze_workbook(content, filename, max_sample_rows)

    elif lower_name.endswith(".xls"):
        return _analyze_xls(content, filename, max_sample_rows)

    elif lower_name.endswith(".txt"):
        return _analyze_txt(content, filename, max_sample_rows)

    elif lower_name.endswith(".json"):
        return _analyze_json(content, filename, max_sample_rows)

    # ==========================================================================
    # VISUAL FORMATS (AI-First via Gemini Vision API)
    # ==========================================================================
    elif lower_name.endswith((".jpg", ".jpeg", ".png")):
        return _analyze_image_with_vision(content, filename)

    elif lower_name.endswith(".pdf"):
        return _analyze_pdf_with_vision(content, filename)

    # ==========================================================================
    # FALLBACK: Detect from content (magic bytes)
    # ==========================================================================
    else:
        # XLSX/ZIP: starts with PK
        if content[:2] == b'PK':
            return analyze_workbook(content, filename, max_sample_rows)
        # XLS (OLE): starts with D0 CF 11 E0
        elif content[:4] == b'\xd0\xcf\x11\xe0':
            return _analyze_xls(content, filename, max_sample_rows)
        # PDF: starts with %PDF
        elif content[:4] == b'%PDF':
            return _analyze_pdf_with_vision(content, filename)
        # PNG: starts with 89 50 4E 47
        elif content[:4] == b'\x89PNG':
            return _analyze_image_with_vision(content, filename)
        # JPEG: starts with FF D8 FF
        elif content[:3] == b'\xff\xd8\xff':
            return _analyze_image_with_vision(content, filename)
        # JSON: starts with [ or {
        elif content[:1] in (b'[', b'{'):
            return _analyze_json(content, filename, max_sample_rows)
        else:
            # Default to CSV/TXT for text-like content
            return _analyze_csv(content, filename, max_sample_rows)


def _analyze_csv(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Analyze a CSV file and return WorkbookAnalysis structure.

    Creates a unified analysis structure compatible with XLSX analysis,
    treating the CSV as a single-sheet workbook.

    Args:
        content: Raw CSV content as bytes
        filename: Original filename
        max_sample_rows: Maximum rows to sample

    Returns:
        WorkbookAnalysis with CSV data as single sheet
    """
    import csv

    reasoning_trace = []

    reasoning_trace.append({
        "type": "thought",
        "content": f"Detectei arquivo CSV: '{filename}'. Vou analisar sua estrutura.",
    })

    # Decode content
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
        reasoning_trace.append({
            "type": "observation",
            "content": "Arquivo usa encoding Latin-1 (não UTF-8)",
        })

    # Detect delimiter
    sample = text[:4096]
    delimiters = [',', ';', '\t', '|']
    delimiter_counts = {d: sample.count(d) for d in delimiters}
    delimiter = max(delimiter_counts, key=delimiter_counts.get)

    reasoning_trace.append({
        "type": "observation",
        "content": f"Delimitador detectado: '{delimiter}' (aparece {delimiter_counts[delimiter]} vezes)",
    })

    # Parse CSV
    lines = text.strip().split('\n')
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)

    if not rows:
        raise ValueError("Arquivo CSV está vazio")

    headers = rows[0]
    data_rows = rows[1:max_sample_rows + 1] if len(rows) > 1 else []
    total_rows = len(rows) - 1  # Exclude header

    reasoning_trace.append({
        "type": "observation",
        "content": f"CSV tem {len(headers)} colunas e {total_rows} linhas de dados",
    })

    # Analyze columns
    columns_analysis = []

    for col_idx, header in enumerate(headers):
        # Get sample values for this column
        sample_values = []
        null_count = 0
        unique_values = set()

        for row in data_rows:
            if col_idx < len(row):
                value = row[col_idx].strip()
                if value:
                    sample_values.append(value)
                    unique_values.add(value)
                else:
                    null_count += 1
            else:
                null_count += 1

        # Detect data type
        data_type = _detect_data_type(sample_values)

        # Normalize column name
        normalized = normalize_column_name(header)

        # Get schema-aware mapping suggestion using the existing function
        suggested_mapping, mapping_confidence = detect_column_mapping(header)

        columns_analysis.append(ColumnAnalysis(
            name=header,
            normalized_name=normalized,
            sample_values=sample_values[:5],
            data_type=data_type,
            unique_count=len(unique_values),
            null_count=null_count,
            is_likely_key=len(unique_values) == len(data_rows) and null_count == 0,
            suggested_mapping=suggested_mapping,
            mapping_confidence=mapping_confidence,
        ))

    # Determine sheet purpose (for CSV, always assume ITEMS)
    purpose = SheetPurpose.ITEMS
    purpose_confidence = 0.8

    reasoning_trace.append({
        "type": "conclusion",
        "content": f"CSV analisado: {len(headers)} colunas, {total_rows} linhas. Pronto para mapeamento.",
    })

    # Create single-sheet analysis
    sheet_analysis = SheetAnalysis(
        name="CSV Data",
        row_count=total_rows,
        column_count=len(headers),
        columns=columns_analysis,
        detected_purpose=purpose,
        purpose_confidence=purpose_confidence,
        has_headers=True,
        suggested_action="process",
        merge_target=None,
        notes=[],
    )

    return WorkbookAnalysis(
        filename=filename,
        sheet_count=1,
        total_rows=total_rows,
        sheets=[sheet_analysis],
        relationships=[],
        recommended_strategy="single_sheet",
        questions_for_user=[],
        reasoning_trace=reasoning_trace,
    )


# =============================================================================
# TXT Analysis (Delimited Text Files)
# =============================================================================


def _analyze_txt(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Analyze delimited text file with automatic delimiter detection.

    Supports delimiters: tab, pipe, semicolon, comma, multiple spaces.
    Internally reuses CSV parsing logic after detecting delimiter.

    Args:
        content: Raw TXT content as bytes
        filename: Original filename
        max_sample_rows: Maximum rows to sample

    Returns:
        WorkbookAnalysis with TXT data as single sheet
    """
    reasoning_trace = []

    reasoning_trace.append({
        "type": "thought",
        "content": f"Detectei arquivo TXT: '{filename}'. Vou identificar o delimitador.",
    })

    # Decode content with fallback encodings
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
            reasoning_trace.append({
                "type": "observation",
                "content": "Arquivo usa encoding Latin-1 (não UTF-8)",
            })
        except UnicodeDecodeError:
            text = content.decode("cp1252", errors="ignore")

    # Detect delimiter - check tab, pipe, semicolon first (more specific)
    sample = text[:4096]
    delimiters = ['\t', '|', ';', ',']
    delimiter_counts = {d: sample.count(d) for d in delimiters}

    # Check for multiple spaces as delimiter
    import re
    multi_space_matches = len(re.findall(r'  +', sample))
    delimiter_counts['  '] = multi_space_matches

    # Choose delimiter with highest count
    delimiter = max(delimiter_counts, key=delimiter_counts.get)

    # Validate delimiter by checking consistency across lines
    lines = text.strip().split('\n')[:10]
    if delimiter and delimiter != '  ':
        counts = [line.count(delimiter) for line in lines if line.strip()]
        if counts and counts[0] > 0 and len(set(counts)) <= 2:
            pass  # Delimiter is consistent
        else:
            delimiter = ','  # Fallback to comma

    delimiter_name = {
        '\t': 'TAB', '|': 'PIPE', ';': 'ponto-e-vírgula',
        ',': 'vírgula', '  ': 'espaços múltiplos'
    }.get(delimiter, delimiter)

    reasoning_trace.append({
        "type": "observation",
        "content": f"Delimitador detectado: {delimiter_name} ({delimiter_counts.get(delimiter, 0)} ocorrências)",
    })

    # Parse with detected delimiter
    import csv
    if delimiter == '  ':
        # Split by multiple spaces
        rows = [re.split(r'  +', line.strip()) for line in text.strip().split('\n') if line.strip()]
    else:
        reader = csv.reader(text.strip().split('\n'), delimiter=delimiter)
        rows = list(reader)

    if not rows:
        raise ValueError("Arquivo TXT está vazio")

    headers = rows[0]
    data_rows = rows[1:max_sample_rows + 1] if len(rows) > 1 else []
    total_rows = len(rows) - 1

    reasoning_trace.append({
        "type": "observation",
        "content": f"TXT tem {len(headers)} colunas e {total_rows} linhas de dados",
    })

    # Analyze columns (same logic as CSV)
    columns_analysis = []
    for col_idx, header in enumerate(headers):
        sample_values = []
        null_count = 0
        unique_values = set()

        for row in data_rows:
            if col_idx < len(row):
                value = row[col_idx].strip() if row[col_idx] else ""
                if value:
                    sample_values.append(value)
                    unique_values.add(value)
                else:
                    null_count += 1
            else:
                null_count += 1

        data_type = _detect_data_type(sample_values)
        suggested_mapping, mapping_confidence = detect_column_mapping(header)

        columns_analysis.append(ColumnAnalysis(
            name=header,
            normalized_name=normalize_column_name(header),
            sample_values=sample_values[:5],
            data_type=data_type,
            unique_count=len(unique_values),
            null_count=null_count,
            is_likely_key=len(unique_values) == len(data_rows) and null_count == 0,
            suggested_mapping=suggested_mapping,
            mapping_confidence=mapping_confidence,
        ))

    reasoning_trace.append({
        "type": "conclusion",
        "content": f"TXT analisado: {len(headers)} colunas, {total_rows} linhas. Pronto para mapeamento.",
    })

    sheet_analysis = SheetAnalysis(
        name="TXT Data",
        row_count=total_rows,
        column_count=len(headers),
        columns=columns_analysis,
        detected_purpose=SheetPurpose.ITEMS,
        purpose_confidence=0.75,
        has_headers=True,
        suggested_action="process",
        merge_target=None,
        notes=[f"Delimitador: {delimiter_name}"],
    )

    return WorkbookAnalysis(
        filename=filename,
        sheet_count=1,
        total_rows=total_rows,
        sheets=[sheet_analysis],
        relationships=[],
        recommended_strategy="single_sheet",
        questions_for_user=[],
        reasoning_trace=reasoning_trace,
    )


# =============================================================================
# JSON Analysis (Array of Objects)
# =============================================================================


def _analyze_json(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Analyze JSON file containing array of objects.

    Expected format: [{"col1": "val1", "col2": "val2"}, ...]
    Each object becomes a row, keys become column headers.

    Args:
        content: Raw JSON content as bytes
        filename: Original filename
        max_sample_rows: Maximum rows to sample

    Returns:
        WorkbookAnalysis with JSON data as single sheet
    """
    import json as json_module

    reasoning_trace = []

    reasoning_trace.append({
        "type": "thought",
        "content": f"Detectei arquivo JSON: '{filename}'. Vou analisar a estrutura.",
    })

    # Decode and parse
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    try:
        data = json_module.loads(text)
    except json_module.JSONDecodeError as e:
        raise ValueError(f"JSON inválido: {str(e)}")

    # Handle different JSON structures
    if isinstance(data, dict):
        # Single object - wrap in array
        if any(isinstance(v, list) for v in data.values()):
            # Nested array inside object - try to find the main data array
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    data = value
                    reasoning_trace.append({
                        "type": "observation",
                        "content": f"JSON contém array aninhado em '{key}', usando como dados principais",
                    })
                    break
            else:
                data = [data]
                reasoning_trace.append({
                    "type": "observation",
                    "content": "JSON contém objeto único, convertido para array",
                })
        else:
            data = [data]
            reasoning_trace.append({
                "type": "observation",
                "content": "JSON contém objeto único, convertido para array",
            })

    if not isinstance(data, list):
        raise ValueError("JSON deve ser um array de objetos ou objeto com array interno")

    if not data:
        raise ValueError("JSON array está vazio")

    # Validate first item is an object
    first_obj = data[0]
    if not isinstance(first_obj, dict):
        raise ValueError("JSON deve conter objetos (dicionários)")

    # Extract headers from all objects (union of all keys)
    all_keys = set()
    for obj in data[:max_sample_rows]:
        if isinstance(obj, dict):
            all_keys.update(obj.keys())
    headers = list(all_keys)
    total_rows = len(data)

    reasoning_trace.append({
        "type": "observation",
        "content": f"JSON tem {len(headers)} colunas (keys) e {total_rows} registros",
    })

    # Analyze columns
    columns_analysis = []
    for header in headers:
        sample_values = []
        null_count = 0
        unique_values = set()

        for obj in data[:max_sample_rows]:
            if isinstance(obj, dict):
                value = obj.get(header)
                if value is not None:
                    str_val = str(value)[:100]
                    sample_values.append(str_val)
                    unique_values.add(str_val)
                else:
                    null_count += 1
            else:
                null_count += 1

        data_type = _detect_data_type(sample_values)
        suggested_mapping, mapping_confidence = detect_column_mapping(header)

        columns_analysis.append(ColumnAnalysis(
            name=header,
            normalized_name=normalize_column_name(header),
            sample_values=sample_values[:5],
            data_type=data_type,
            unique_count=len(unique_values),
            null_count=null_count,
            is_likely_key=len(unique_values) == min(len(data), max_sample_rows) and null_count == 0,
            suggested_mapping=suggested_mapping,
            mapping_confidence=mapping_confidence,
        ))

    reasoning_trace.append({
        "type": "conclusion",
        "content": f"JSON analisado: {len(headers)} campos, {total_rows} registros. Pronto para mapeamento.",
    })

    sheet_analysis = SheetAnalysis(
        name="JSON Data",
        row_count=total_rows,
        column_count=len(headers),
        columns=columns_analysis,
        detected_purpose=SheetPurpose.ITEMS,
        purpose_confidence=0.80,
        has_headers=True,
        suggested_action="process",
        merge_target=None,
        notes=[],
    )

    return WorkbookAnalysis(
        filename=filename,
        sheet_count=1,
        total_rows=total_rows,
        sheets=[sheet_analysis],
        relationships=[],
        recommended_strategy="single_sheet",
        questions_for_user=[],
        reasoning_trace=reasoning_trace,
    )


# =============================================================================
# XLS Analysis (Excel 97-2003 Format)
# =============================================================================

# Lazy import for xlrd (cold start optimization)
_xlrd_module = None


def _get_xlrd():
    """Lazy load xlrd for XLS support (cold start optimization)."""
    global _xlrd_module
    if _xlrd_module is None:
        try:
            import xlrd
            _xlrd_module = xlrd
            logger.info("[SheetAnalyzer] xlrd loaded for XLS support")
        except ImportError:
            raise ImportError(
                "xlrd package required for XLS (Excel 97-2003) files. "
                "Add 'xlrd>=2.0.1' to requirements.txt"
            )
    return _xlrd_module


def _analyze_xls(
    content: bytes,
    filename: str,
    max_sample_rows: int = 20,
) -> WorkbookAnalysis:
    """
    Analyze XLS (Excel 97-2003) file using xlrd.

    Mirrors analyze_workbook() pattern but with xlrd API.

    Args:
        content: Raw XLS content as bytes
        filename: Original filename
        max_sample_rows: Maximum rows to sample per sheet

    Returns:
        WorkbookAnalysis with XLS data
    """
    xlrd = _get_xlrd()

    reasoning_trace = []

    reasoning_trace.append({
        "type": "thought",
        "content": f"Detectei arquivo XLS (Excel 97-2003): '{filename}'",
    })

    # Open workbook
    try:
        wb = xlrd.open_workbook(file_contents=content)
    except xlrd.XLRDError as e:
        raise ValueError(f"Erro ao abrir arquivo XLS: {str(e)}")

    reasoning_trace.append({
        "type": "observation",
        "content": f"Arquivo tem {wb.nsheets} aba(s): {', '.join(wb.sheet_names())}",
    })

    sheets_analysis = []
    total_rows = 0

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name

        reasoning_trace.append({
            "type": "action",
            "content": f"Analisando aba '{sheet_name}'",
        })

        if ws.nrows == 0:
            reasoning_trace.append({
                "type": "observation",
                "content": f"Aba '{sheet_name}' está vazia, pulando",
            })
            continue

        # Get headers from first row
        headers = []
        for c in range(ws.ncols):
            cell_value = ws.cell_value(0, c)
            headers.append(str(cell_value) if cell_value else f"Column_{c}")

        row_count = ws.nrows - 1  # Exclude header

        # Sample data rows
        sample_rows = []
        for r in range(1, min(max_sample_rows + 1, ws.nrows)):
            row_values = [ws.cell_value(r, c) for c in range(ws.ncols)]
            sample_rows.append(row_values)

        # Analyze columns
        columns = []
        for col_idx, header in enumerate(headers):
            col_values = [
                row[col_idx] if col_idx < len(row) else None
                for row in sample_rows
            ]

            sample_values = [str(v) for v in col_values[:5] if v is not None and str(v).strip()]
            data_type = detect_data_type(col_values)
            mapping, confidence = detect_column_mapping(header)

            unique_vals = set(str(v) for v in col_values if v is not None and str(v).strip())

            columns.append(ColumnAnalysis(
                name=header,
                normalized_name=normalize_column_name(header),
                sample_values=sample_values,
                data_type=data_type,
                unique_count=len(unique_vals),
                null_count=sum(1 for v in col_values if v is None or str(v).strip() == ""),
                is_likely_key=len(unique_vals) == len([v for v in col_values if v]) and len(unique_vals) > 1,
                suggested_mapping=mapping,
                mapping_confidence=confidence,
            ))

        # Detect sheet purpose
        purpose, purpose_conf = detect_sheet_purpose(sheet_name, columns, row_count)

        # Determine action
        if purpose == SheetPurpose.SUMMARY or purpose == SheetPurpose.METADATA:
            action = "skip"
        elif purpose == SheetPurpose.SERIALS:
            action = "merge_with"
        else:
            action = "process"

        sheets_analysis.append(SheetAnalysis(
            name=sheet_name,
            row_count=row_count,
            column_count=len(headers),
            columns=columns,
            detected_purpose=purpose,
            purpose_confidence=purpose_conf,
            has_headers=True,
            suggested_action=action,
            merge_target=None,
            notes=[],
        ))

        total_rows += row_count

        reasoning_trace.append({
            "type": "observation",
            "content": (
                f"Aba '{sheet_name}': {row_count} linhas, "
                f"{len(headers)} colunas, propósito: {purpose.value}"
            ),
        })

    if not sheets_analysis:
        raise ValueError("Arquivo XLS não contém dados válidos")

    # Detect relationships between sheets
    relationships = detect_sheet_relationships(sheets_analysis)
    strategy = _determine_processing_strategy(sheets_analysis, relationships)

    reasoning_trace.append({
        "type": "conclusion",
        "content": f"XLS analisado: {len(sheets_analysis)} aba(s), {total_rows} linhas. Estratégia: {strategy}",
    })

    analysis = WorkbookAnalysis(
        filename=filename,
        sheet_count=len(sheets_analysis),
        total_rows=total_rows,
        sheets=sheets_analysis,
        relationships=relationships,
        recommended_strategy=strategy,
        questions_for_user=[],
        reasoning_trace=reasoning_trace,
    )

    # Generate questions
    analysis.questions_for_user = generate_questions(analysis)

    return analysis


# =============================================================================
# Vision-Based Analysis (AI-First via Gemini Vision API)
# =============================================================================


def _analyze_image_with_vision(
    content: bytes,
    filename: str,
) -> WorkbookAnalysis:
    """
    Analyze image containing tabular data using Gemini Vision API.

    This is the AI-First approach - all visual data extraction is done
    by the LLM, not by traditional OCR libraries.

    Supports: JPG, JPEG, PNG, GIF

    Args:
        content: Raw image content as bytes
        filename: Original filename

    Returns:
        WorkbookAnalysis with extracted table data
    """
    # Lazy import of vision extractor
    from tools.vision_table_extractor import extract_table_from_image

    return extract_table_from_image(content, filename)


def _analyze_pdf_with_vision(
    content: bytes,
    filename: str,
) -> WorkbookAnalysis:
    """
    Analyze PDF document using Gemini Vision API.

    This is the AI-First approach - PDF is processed directly by
    Gemini Vision, not by traditional PDF libraries like PyPDF2.

    Args:
        content: Raw PDF content as bytes
        filename: Original filename

    Returns:
        WorkbookAnalysis with extracted table data
    """
    # Lazy import of vision extractor
    from tools.vision_table_extractor import extract_table_from_pdf

    return extract_table_from_pdf(content, filename)


def _detect_data_type(values: List[str]) -> str:
    """Detect data type from sample values."""
    if not values:
        return "text"

    numeric_count = 0
    date_count = 0

    for v in values:
        # Check numeric
        try:
            float(v.replace(",", ".").replace(" ", ""))
            numeric_count += 1
            continue
        except ValueError:
            pass

        # Check date patterns
        if any(sep in v for sep in ["/", "-"]) and len(v) <= 20:
            parts = v.replace("-", "/").split("/")
            if len(parts) >= 2 and all(p.isdigit() for p in parts if p):
                date_count += 1

    if numeric_count >= len(values) * 0.8:
        return "number"
    elif date_count >= len(values) * 0.8:
        return "date"
    else:
        return "text"


def load_workbook_smart(content: bytes):
    """
    Load workbook content smartly, detecting format.

    Returns an object that can be used with both pandas ExcelFile
    or openpyxl Workbook APIs.

    Args:
        content: Raw file content

    Returns:
        Workbook object (openpyxl or pandas ExcelFile)
    """
    # Check if content is XLSX (starts with PK = ZIP signature)
    if content[:2] == b'PK':
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    else:
        # For CSV, return a simple wrapper
        return _CSVWorkbookWrapper(content)


class _CSVWorkbookWrapper:
    """Wrapper to make CSV content look like a workbook for compatibility."""

    def __init__(self, content: bytes):
        self.content = content
        self.sheetnames = ["CSV Data"]
        self._data = None

    def __getitem__(self, sheet_name: str):
        return _CSVSheetWrapper(self.content)

    def close(self):
        pass


class _CSVSheetWrapper:
    """Wrapper to make CSV sheet look like openpyxl worksheet."""

    def __init__(self, content: bytes):
        import csv as csv_module

        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        # Detect delimiter
        sample = text[:4096]
        delimiters = [',', ';', '\t', '|']
        delimiter_counts = {d: sample.count(d) for d in delimiters}
        delimiter = max(delimiter_counts, key=delimiter_counts.get)

        lines = text.strip().split('\n')
        reader = csv_module.reader(lines, delimiter=delimiter)
        self._rows = list(reader)

    def iter_rows(self, min_row: int = 1, max_row: int = None, values_only: bool = False):
        """Iterate rows like openpyxl."""
        start_idx = min_row - 1
        end_idx = max_row if max_row else len(self._rows)

        for row in self._rows[start_idx:end_idx]:
            yield tuple(row)
