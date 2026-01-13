# =============================================================================
# Gemini Text Analyzer - AI-First File Analysis for CSV/XLSX
# =============================================================================
# Analyzes text-based files (CSV, XLSX) using Gemini Pro.
# This is the AI-First approach - the LLM understands the data semantically
# and provides column mappings, confidence scores, and HIL questions.
#
# Philosophy: OBSERVE (Text) -> THINK (LLM) -> LEARN -> ACT
#
# For CSV: Read raw text -> Send to Gemini
# For XLSX: Extract text (openpyxl) -> Send to Gemini
#
# Module: Gestao de Ativos -> Gestao de Estoque -> Smart Import
# Author: Faiston NEXO Team
# Created: January 2026
# =============================================================================

import logging
import json
import io
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

# =============================================================================
# Lazy Import for Cold Start Optimization
# =============================================================================

_genai_client = None
_s3_client = None

S3_BUCKET = "faiston-one-sga-documents-prod"
AWS_REGION = "us-east-2"


def _get_genai_client():
    """
    Lazy initialization of Gemini client.

    Cold start optimization: Only import google.genai when needed.
    This prevents the 30-second AgentCore timeout.
    """
    global _genai_client
    if _genai_client is None:
        from google import genai
        _genai_client = genai.Client()
        logger.info("[GeminiTextAnalyzer] Gemini client initialized")
    return _genai_client


def _get_s3_client():
    """Lazy load S3 client for cold start optimization."""
    global _s3_client
    if _s3_client is None:
        import boto3
        _s3_client = boto3.client("s3", region_name=AWS_REGION)
        logger.info("[GeminiTextAnalyzer] S3 client initialized")
    return _s3_client


# =============================================================================
# Portuguese Analysis Prompt for Inventory Data
# =============================================================================

INVENTORY_ANALYSIS_PROMPT = """Voce e um especialista em analise de dados de inventario e estoque com comportamento AGI-like.

## TAREFA
Analise o conteudo do arquivo (CSV ou planilha) e:
1. Identifique TODAS as colunas e seus tipos de dados
2. Sugira mapeamento para o schema do banco de dados PostgreSQL
3. Calcule confianca do mapeamento (0.0 a 1.0)
4. Gere perguntas para o usuario se confianca < 0.80
5. IDENTIFIQUE colunas que NAO existem no schema do DB (unmapped_columns)
6. Se houver respostas do usuario, RE-ANALISE ajustando seus mapeamentos

## SCHEMA DO BANCO DE DADOS POSTGRESQL
{schema_context}

## PADROES APRENDIDOS (memoria do agente)
{memory_context}

## RESPOSTAS DO USUARIO (HIL - Human-in-the-Loop)
{user_responses}

## COMENTARIOS/INSTRUCOES DO USUARIO
{user_comments}

## CAMPOS COMUNS DE INVENTARIO (procure ativamente)
- Part Number / Codigo / SKU / Material / PN / Cod
- Descricao / Nome do Item / Produto / Description
- Quantidade / Qtd / Qty / Quant
- Numero de Serie / Serial / SN / NS / IMEI / S/N
- Local / Deposito / Armazem / Localizacao / Location
- Projeto / Obra / Cliente / Project
- Fornecedor / Fabricante / Marca / Supplier
- Valor / Preco / Custo / Unit Value / Price
- NCM / Codigo Fiscal
- Data / Data de Entrada / Validade / Date
- Status / Situacao / Condicao / State

## FORMATO DE RESPOSTA (JSON OBRIGATORIO)

Retorne APENAS JSON valido, sem markdown, sem explicacoes:

{{
  "success": true,
  "file_type": "csv",
  "analysis_confidence": 0.85,
  "analysis_round": 1,
  "quality_issues": [],
  "detected_encoding": "utf-8",
  "detected_delimiter": ";",
  "row_count": 1688,
  "column_count": 10,
  "columns": [
    {{
      "source_name": "Nome Original da Coluna",
      "normalized_name": "nome_normalizado",
      "suggested_mapping": "part_number",
      "mapping_confidence": 0.95,
      "data_type": "string",
      "sample_values": ["valor1", "valor2", "valor3"],
      "is_required": true,
      "is_unmapped": false
    }}
  ],
  "suggested_mappings": {{
    "Nome Original": "target_field",
    "Outra Coluna": "another_field"
  }},
  "unmapped_columns": [
    {{
      "source_name": "N TICKET",
      "reason": "Coluna nao existe no schema do banco de dados",
      "suggested_action": "metadata",
      "description": "Parece ser numero de chamado/ticket de suporte"
    }}
  ],
  "hil_questions": [
    {{
      "id": "q1",
      "field": "nome_da_coluna",
      "question": "Esta coluna contem numeros de serie ou codigos de lote?",
      "options": ["Numero de Serie", "Codigo de Lote", "Outro"],
      "reason": "Ambiguidade entre serial e lote",
      "priority": "high"
    }}
  ],
  "unmapped_questions": [
    {{
      "id": "uq1",
      "field": "N TICKET",
      "question": "A coluna 'N TICKET' nao existe no banco de dados. O que deseja fazer?",
      "options": ["Ignorar (dados serao perdidos)", "Guardar em metadata (preservar)", "Solicitar criacao de campo no DB"],
      "reason": "Coluna nao mapeada - requer decisao do usuario"
    }}
  ],
  "all_questions_answered": false,
  "ready_for_import": false,
  "recommended_action": "needs_user_input",
  "notes": "Arquivo com 1688 registros de expedicao. 4 colunas nao mapeadas."
}}

## REGRAS IMPORTANTES

1. **Mapeamento automatico**:
   - Se confianca >= 0.80: mapeamento automatico permitido
   - Se confianca < 0.80: OBRIGATORIO gerar pergunta HIL

2. **Campos obrigatorios para importacao**:
   - part_number (codigo do produto)
   - quantity (quantidade) - OU pode ser calculado se houver serial_number

3. **Valores especiais**:
   - Celulas vazias: null
   - Valores numericos: manter como number
   - Datas: preservar formato original

4. **Confianca do mapeamento**:
   - 0.95+: Match exato com nome do campo
   - 0.80-0.94: Match por sinonimo ou padrao aprendido
   - 0.60-0.79: Inferencia por amostra de dados
   - <0.60: Incerto - requer HIL

5. **recommended_action**:
   - "ready_for_import": Todos mapeamentos com confianca >= 0.80 E todas perguntas respondidas
   - "needs_user_input": Algum mapeamento < 0.80 OU colunas nao mapeadas sem decisao
   - "error": Arquivo invalido ou campos obrigatorios faltando

6. **Colunas NAO MAPEADAS (unmapped_columns) - CRITICO**:
   - Se coluna do arquivo NAO existe no schema PostgreSQL: OBRIGATORIO adicionar em unmapped_columns
   - OBRIGATORIO gerar pergunta em unmapped_questions com 3 opcoes:
     a) Ignorar (dados serao perdidos)
     b) Guardar em metadata (preservar em campo JSONB)
     c) Solicitar criacao de campo no DB (usuario deve contatar equipe de TI)
   - Import NAO pode prosseguir ate usuario decidir sobre TODAS colunas nao mapeadas

7. **Re-analise com respostas do usuario (AGI-like)**:
   - Se user_responses NAO esta vazio: re-analisar mapeamentos considerando as respostas
   - Ajustar mapeamentos e confidences baseado no feedback do usuario
   - Se usuario disse "Sim" para uma pergunta: aumentar confidence para 1.0
   - Continuar gerando perguntas ate TODAS terem confidence >= 0.80

8. **Calculo de quantidade (quantity)**:
   - Se coluna quantity NAO existe mas serial_number existe:
     - Agrupar por part_number
     - Contar serial_numbers unicos = quantity
     - Cada part_number deve ser UNICO no resultado final
   - Se usuario instruiu via comentarios: seguir instrucao

9. **all_questions_answered**:
   - true: Todas perguntas HIL e unmapped foram respondidas
   - false: Ainda ha perguntas pendentes

10. **ready_for_import**:
    - true: Condicoes para import: all_questions_answered=true E analysis_confidence >= 0.80
    - false: Ainda precisa de input do usuario

## CONTEUDO DO ARQUIVO PARA ANALISE
{file_content}

Analise e retorne o JSON:"""


# =============================================================================
# Default Schema Context (PostgreSQL pending_entry_items)
# =============================================================================

DEFAULT_SCHEMA_CONTEXT = """
Tabela: pending_entry_items (itens pendentes de entrada no estoque)

Campos:
- id: UUID (auto-gerado)
- part_number: VARCHAR(100) - Codigo do produto (OBRIGATORIO)
- description: TEXT - Descricao do item
- quantity: INTEGER - Quantidade (OBRIGATORIO)
- serial_number: VARCHAR(100) - Numero de serie (opcional)
- location_code: VARCHAR(50) - Codigo do local de destino
- project_code: VARCHAR(50) - Codigo do projeto/obra
- supplier: VARCHAR(200) - Fornecedor
- unit_value: DECIMAL(12,2) - Valor unitario
- nf_number: VARCHAR(50) - Numero da nota fiscal
- nf_date: DATE - Data da nota fiscal
- status: VARCHAR(20) - Status (pending, processed, error)
- created_at: TIMESTAMP - Data de criacao
- created_by: VARCHAR(100) - Usuario que criou
- metadata: JSONB - Campos extras dinamicos
"""


# =============================================================================
# File Content Extraction
# =============================================================================

def _extract_csv_content(content: bytes, max_rows: int = 50) -> str:
    """
    Extract CSV content as text for Gemini analysis.

    Args:
        content: Raw CSV bytes
        max_rows: Max rows to sample (for prompt efficiency)

    Returns:
        CSV text with headers + sample rows
    """
    # Detect encoding
    try:
        text = content.decode("utf-8")
        encoding = "utf-8"
    except UnicodeDecodeError:
        text = content.decode("latin-1")
        encoding = "latin-1"

    lines = text.strip().split('\n')

    # Take header + sample rows
    if len(lines) > max_rows + 1:
        sample_lines = lines[:1] + lines[1:max_rows + 1]
        sample_text = '\n'.join(sample_lines)
        sample_text += f"\n\n[... mais {len(lines) - max_rows - 1} linhas ...]"
    else:
        sample_text = text

    return f"Encoding: {encoding}\nTotal linhas: {len(lines)}\n\n{sample_text}"


def _extract_xlsx_content(content: bytes, max_rows: int = 50) -> str:
    """
    Extract XLSX content as text for Gemini analysis.

    Args:
        content: Raw XLSX bytes
        max_rows: Max rows to sample

    Returns:
        JSON-like text representation of the spreadsheet
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    sheets_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            continue

        headers = [str(h) if h else f"Coluna_{i}" for i, h in enumerate(headers)]

        rows = []
        for i, row_values in enumerate(rows_iter):
            if i >= max_rows:
                break
            row_dict = {}
            for j, val in enumerate(row_values):
                if j < len(headers):
                    row_dict[headers[j]] = str(val) if val is not None else None
            rows.append(row_dict)

        total_rows = ws.max_row - 1 if ws.max_row else 0

        sheets_data.append({
            "sheet_name": sheet_name,
            "headers": headers,
            "sample_rows": rows,
            "total_rows": total_rows,
        })

    wb.close()

    return json.dumps(sheets_data, indent=2, ensure_ascii=False, default=str)


def _extract_xls_content(content: bytes, max_rows: int = 50) -> str:
    """
    Extract XLS (legacy Excel) content as text.

    Args:
        content: Raw XLS bytes
        max_rows: Max rows to sample

    Returns:
        JSON-like text representation
    """
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)

    sheets_data = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)

        if ws.nrows == 0:
            continue

        headers = [str(ws.cell_value(0, c)) or f"Coluna_{c}" for c in range(ws.ncols)]

        rows = []
        for r in range(1, min(ws.nrows, max_rows + 1)):
            row_dict = {}
            for c in range(ws.ncols):
                val = ws.cell_value(r, c)
                row_dict[headers[c]] = str(val) if val else None
            rows.append(row_dict)

        sheets_data.append({
            "sheet_name": ws.name,
            "headers": headers,
            "sample_rows": rows,
            "total_rows": ws.nrows - 1,
        })

    return json.dumps(sheets_data, indent=2, ensure_ascii=False, default=str)


# =============================================================================
# Response Parsing
# =============================================================================

def _extract_json_from_response(response_text: str) -> Optional[Dict]:
    """
    Extract JSON from Gemini response text.

    Handles cases where JSON is wrapped in markdown code blocks.
    """
    text = response_text.strip()

    # Try direct JSON parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try to extract from markdown code block
    if "```json" in text:
        start = text.find("```json") + 7
        end = text.find("```", start)
        if end > start:
            try:
                return json.loads(text[start:end].strip())
            except json.JSONDecodeError:
                pass

    # Try to find JSON object anywhere
    start_idx = text.find("{")
    if start_idx >= 0:
        brace_count = 0
        for i, char in enumerate(text[start_idx:], start_idx):
            if char == "{":
                brace_count += 1
            elif char == "}":
                brace_count -= 1
                if brace_count == 0:
                    try:
                        return json.loads(text[start_idx:i + 1])
                    except json.JSONDecodeError:
                        pass
                    break

    return None


# =============================================================================
# User Response Formatting (AGI-Like Multi-Round HIL)
# =============================================================================

def _format_user_responses(user_responses: List[Dict[str, Any]]) -> str:
    """
    Format user responses for inclusion in the Gemini prompt.

    This enables the AGI-like behavior where user responses feed back
    into the LLM for re-analysis.

    Args:
        user_responses: List of responses from previous HIL rounds
                        Format: [{"question_id": "q1", "answer": "Numero de Serie"}]

    Returns:
        Formatted string for the prompt
    """
    if not user_responses:
        return "Nenhuma resposta do usuario ainda (primeira analise)."

    lines = ["Respostas do usuario das rodadas anteriores:"]
    for i, resp in enumerate(user_responses, 1):
        q_id = resp.get("question_id", f"q{i}")
        field = resp.get("field", "unknown")
        answer = resp.get("answer", "")
        lines.append(f"  {i}. [{q_id}] Campo '{field}': {answer}")

    lines.append("")
    lines.append("IMPORTANTE: Re-analise os mapeamentos considerando estas respostas!")
    lines.append("Se o usuario confirmou um mapeamento, aumente a confianca para 1.0")

    return "\n".join(lines)


# =============================================================================
# Main Analysis Functions
# =============================================================================

async def analyze_file_with_gemini(
    s3_key: str,
    schema_context: str = None,
    memory_context: str = None,
    user_responses: List[Dict[str, Any]] = None,
    user_comments: str = None,
    analysis_round: int = 1,
) -> Dict[str, Any]:
    """
    Analyze file from S3 using Gemini Pro (AI-First with AGI-like behavior).

    This is the main entry point for AI-First file analysis with iterative HIL.

    Flow (Multi-Round AGI Loop):
    1. Download file from S3
    2. Extract text content (CSV as-is, XLSX to JSON)
    3. Send to Gemini with schema + memory + user_responses + user_comments
    4. Return analysis with mappings, confidence, and questions
    5. If user responds, call again with responses for RE-ANALYSIS

    Args:
        s3_key: S3 key where file is stored
        schema_context: Target PostgreSQL schema description
        memory_context: Prior learned patterns from AgentCore Memory
        user_responses: Accumulated HIL responses from previous rounds
                        Format: [{"question_id": "q1", "answer": "Numero de Serie"}]
        user_comments: Free-text instructions/feedback from user
        analysis_round: Current round number (1 = first analysis, 2+ = re-analysis)

    Returns:
        {
            "success": bool,
            "file_type": str,
            "analysis_confidence": float,
            "analysis_round": int,
            "columns": [...],
            "suggested_mappings": {...},
            "unmapped_columns": [...],
            "hil_questions": [...],
            "unmapped_questions": [...],
            "all_questions_answered": bool,
            "ready_for_import": bool,
            "recommended_action": str,
        }
    """
    logger.info(f"[GeminiTextAnalyzer] Analyzing file: {s3_key} (Round {analysis_round})")
    if user_responses:
        logger.info(f"[GeminiTextAnalyzer] Re-analysis with {len(user_responses)} user responses")

    try:
        # 1. Download file from S3
        import unicodedata

        s3 = _get_s3_client()
        # NFC normalize S3 key to match how files were uploaded
        # Prevents NoSuchKey errors with Portuguese characters (Ç, Ã, Õ)
        normalized_key = unicodedata.normalize("NFC", s3_key)
        response = s3.get_object(Bucket=S3_BUCKET, Key=normalized_key)
        content = response["Body"].read()

        filename = s3_key.split("/")[-1] if "/" in s3_key else s3_key
        filename_lower = filename.lower()

        # 2. Extract text content based on file type
        if filename_lower.endswith(".csv"):
            file_content = _extract_csv_content(content)
            file_type = "csv"
        elif filename_lower.endswith(".xlsx"):
            file_content = _extract_xlsx_content(content)
            file_type = "xlsx"
        elif filename_lower.endswith(".xls"):
            file_content = _extract_xls_content(content)
            file_type = "xls"
        else:
            return {
                "success": False,
                "error": f"Formato nao suportado: {filename}",
                "file_type": "unknown",
            }

        # 3. Build prompt with context (AGI-like: includes user responses)
        user_responses_text = _format_user_responses(user_responses)
        user_comments_text = user_comments or "Nenhum comentario adicional."

        prompt = INVENTORY_ANALYSIS_PROMPT.format(
            schema_context=schema_context or DEFAULT_SCHEMA_CONTEXT,
            memory_context=memory_context or "Nenhum padrao aprendido ainda.",
            user_responses=user_responses_text,
            user_comments=user_comments_text,
            file_content=file_content,
        )

        # 4. Call Gemini Pro
        client = _get_genai_client()
        response = client.models.generate_content(
            model="gemini-2.5-pro",
            contents=prompt,
            config={
                "response_mime_type": "application/json",
            }
        )

        # 5. Parse response
        result = _extract_json_from_response(response.text)

        if not result:
            logger.error(f"[GeminiTextAnalyzer] Failed to parse response: {response.text[:500]}")
            return {
                "success": False,
                "error": "Falha ao processar resposta do Gemini",
                "raw_response": response.text[:1000],
            }

        # 6. Add metadata
        result["filename"] = filename
        result["file_type"] = file_type
        result["s3_key"] = s3_key
        result["analysis_round"] = analysis_round

        # Ensure AGI-like fields exist
        if "unmapped_columns" not in result:
            result["unmapped_columns"] = []
        if "unmapped_questions" not in result:
            result["unmapped_questions"] = []
        if "all_questions_answered" not in result:
            # Check if all questions are answered
            hil_questions = result.get("hil_questions", [])
            unmapped_questions = result.get("unmapped_questions", [])
            total_questions = len(hil_questions) + len(unmapped_questions)
            result["all_questions_answered"] = total_questions == 0
        if "ready_for_import" not in result:
            result["ready_for_import"] = (
                result.get("all_questions_answered", False) and
                result.get("analysis_confidence", 0) >= 0.80
            )

        logger.info(
            f"[GeminiTextAnalyzer] Round {analysis_round} complete: "
            f"confidence={result.get('analysis_confidence', 0):.2f}, "
            f"ready={result.get('ready_for_import', False)}, "
            f"questions={len(result.get('hil_questions', []))}, "
            f"unmapped={len(result.get('unmapped_columns', []))}"
        )

        return result

    except Exception as e:
        logger.error(f"[GeminiTextAnalyzer] Error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
        }


async def extract_data_with_gemini(
    s3_key: str,
    column_mappings: Dict[str, str],
    target_schema: str = None,
    max_rows: int = 5000,
) -> Dict[str, Any]:
    """
    Extract and transform data from file using Gemini (AI-First).

    This function reads the file and uses Gemini to extract
    rows transformed according to the column mappings.

    Args:
        s3_key: S3 key where file is stored
        column_mappings: Validated mappings {source_column: target_field}
        target_schema: Target table schema
        max_rows: Maximum rows to extract

    Returns:
        {
            "success": bool,
            "rows": [...],
            "row_count": int,
            "errors": [...],
        }
    """
    logger.info(f"[GeminiTextAnalyzer] Extracting data from: {s3_key}")

    try:
        # For extraction, we need the full data (not just sample)
        # Use traditional parsing here since Gemini has context limits
        # The AI-First part was the ANALYSIS, extraction is mechanical
        import unicodedata

        s3 = _get_s3_client()
        # NFC normalize S3 key to match how files were uploaded
        # Prevents NoSuchKey errors with Portuguese characters (Ç, Ã, Õ)
        normalized_key = unicodedata.normalize("NFC", s3_key)
        response = s3.get_object(Bucket=S3_BUCKET, Key=normalized_key)
        content = response["Body"].read()

        filename = s3_key.split("/")[-1] if "/" in s3_key else s3_key
        filename_lower = filename.lower()

        rows = []
        errors = []

        if filename_lower.endswith(".csv"):
            rows, errors = _extract_csv_rows(content, column_mappings, max_rows)
        elif filename_lower.endswith(".xlsx"):
            rows, errors = _extract_xlsx_rows(content, column_mappings, max_rows)
        elif filename_lower.endswith(".xls"):
            rows, errors = _extract_xls_rows(content, column_mappings, max_rows)
        else:
            return {
                "success": False,
                "error": f"Formato nao suportado: {filename}",
                "rows": [],
            }

        logger.info(f"[GeminiTextAnalyzer] Extracted {len(rows)} rows, {len(errors)} errors")

        return {
            "success": True,
            "rows": rows,
            "row_count": len(rows),
            "errors": errors[:50],  # Limit errors
            "errors_count": len(errors),
        }

    except Exception as e:
        logger.error(f"[GeminiTextAnalyzer] Extract error: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "rows": [],
        }


# =============================================================================
# Data Extraction Helpers (Mechanical, not AI)
# =============================================================================

def _extract_csv_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform CSV rows using mappings."""
    import csv as csv_module

    # Detect encoding
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    # Detect delimiter
    sample = text[:4096]
    delimiters = [',', ';', '\t', '|']
    delimiter_counts = {d: sample.count(d) for d in delimiters}
    delimiter = max(delimiter_counts, key=delimiter_counts.get)

    lines = text.strip().split('\n')
    reader = csv_module.DictReader(lines, delimiter=delimiter)

    rows = []
    errors = []

    for i, row in enumerate(reader):
        if i >= max_rows:
            break

        try:
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row:
                    transformed[target_field] = row[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})

    return rows, errors


def _extract_xlsx_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform XLSX rows using mappings."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)

    if not headers:
        wb.close()
        return [], [{"row": 1, "error": "No headers found"}]

    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

    rows = []
    errors = []

    for i, row_values in enumerate(rows_iter):
        if i >= max_rows:
            break

        try:
            # Build row dict
            row_dict = {}
            for j, val in enumerate(row_values):
                if j < len(headers):
                    row_dict[headers[j]] = val

            # Transform using mappings
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row_dict:
                    transformed[target_field] = row_dict[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": i + 2, "error": str(e)})

    wb.close()
    return rows, errors


def _extract_xls_rows(
    content: bytes,
    column_mappings: Dict[str, str],
    max_rows: int,
) -> tuple:
    """Extract and transform XLS rows using mappings."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        return [], [{"row": 1, "error": "Empty sheet"}]

    headers = [str(ws.cell_value(0, c)) or f"Column_{c}" for c in range(ws.ncols)]

    rows = []
    errors = []

    for r in range(1, min(ws.nrows, max_rows + 1)):
        try:
            # Build row dict
            row_dict = {}
            for c in range(ws.ncols):
                row_dict[headers[c]] = ws.cell_value(r, c)

            # Transform using mappings
            transformed = {}
            for source_col, target_field in column_mappings.items():
                if source_col in row_dict:
                    transformed[target_field] = row_dict[source_col]

            if transformed:
                rows.append(transformed)
        except Exception as e:
            errors.append({"row": r + 1, "error": str(e)})

    return rows, errors


# =============================================================================
# Convenience Functions
# =============================================================================

def get_default_schema_context() -> str:
    """Get the default PostgreSQL schema context."""
    return DEFAULT_SCHEMA_CONTEXT


async def analyze_file_simple(s3_key: str) -> Dict[str, Any]:
    """
    Simple wrapper for file analysis with default context (Round 1).

    Use this for quick analysis without custom schema/memory.
    """
    return await analyze_file_with_gemini(
        s3_key=s3_key,
        schema_context=DEFAULT_SCHEMA_CONTEXT,
        memory_context=None,
        user_responses=None,
        user_comments=None,
        analysis_round=1,
    )


async def re_analyze_with_responses(
    s3_key: str,
    user_responses: List[Dict[str, Any]],
    user_comments: str = None,
    schema_context: str = None,
    memory_context: str = None,
    previous_round: int = 1,
) -> Dict[str, Any]:
    """
    Re-analyze file with user responses (AGI-like multi-round).

    Use this to continue the HIL dialogue after user responds to questions.

    Args:
        s3_key: S3 key where file is stored
        user_responses: User's answers to HIL questions
                        Format: [{"question_id": "q1", "field": "SERIAL", "answer": "Numero de Serie"}]
        user_comments: Free-text instructions from user
        schema_context: PostgreSQL schema
        memory_context: Learned patterns
        previous_round: Previous round number

    Returns:
        Updated analysis with adjusted mappings and potentially new questions
    """
    return await analyze_file_with_gemini(
        s3_key=s3_key,
        schema_context=schema_context or DEFAULT_SCHEMA_CONTEXT,
        memory_context=memory_context,
        user_responses=user_responses,
        user_comments=user_comments,
        analysis_round=previous_round + 1,
    )


def check_analysis_ready(analysis_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Check if analysis is ready for import or needs more HIL.

    Args:
        analysis_result: Result from analyze_file_with_gemini

    Returns:
        {
            "ready": bool,
            "reason": str,
            "pending_questions": int,
            "unmapped_columns": int,
            "confidence": float,
        }
    """
    hil_questions = analysis_result.get("hil_questions", [])
    unmapped_questions = analysis_result.get("unmapped_questions", [])
    unmapped_columns = analysis_result.get("unmapped_columns", [])
    confidence = analysis_result.get("analysis_confidence", 0)
    ready = analysis_result.get("ready_for_import", False)

    pending = len(hil_questions) + len(unmapped_questions)

    if ready:
        reason = "Analise completa - pronto para importacao"
    elif pending > 0:
        reason = f"{pending} pergunta(s) pendente(s)"
    elif len(unmapped_columns) > 0:
        reason = f"{len(unmapped_columns)} coluna(s) nao mapeada(s) sem decisao"
    elif confidence < 0.80:
        reason = f"Confianca baixa ({confidence:.0%})"
    else:
        reason = "Precisa de aprovacao do usuario"

    return {
        "ready": ready,
        "reason": reason,
        "pending_questions": pending,
        "unmapped_columns": len(unmapped_columns),
        "confidence": confidence,
        "round": analysis_result.get("analysis_round", 1),
    }
