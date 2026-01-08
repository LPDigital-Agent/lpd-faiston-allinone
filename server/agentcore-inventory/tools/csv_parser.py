# =============================================================================
# CSV/Excel Parser Tool - SGA Inventory Module
# =============================================================================
# Parses CSV and Excel files for bulk import.
# Auto-detects delimiters and maps columns to expected fields.
#
# SCHEMA-AWARE: Now uses SchemaColumnMatcher for dynamic column matching
# against PostgreSQL schema instead of hardcoded patterns.
#
# Author: Faiston NEXO Team
# Updated: January 2026 - Schema-aware matching
# =============================================================================

import csv
import io
import logging
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from enum import Enum

logger = logging.getLogger(__name__)


# =============================================================================
# Types and Constants
# =============================================================================

class ImportFileType(Enum):
    """Supported file types for import."""
    CSV = "csv"
    XLSX = "xlsx"


@dataclass
class ColumnMapping:
    """Mapping between file column and expected field."""
    file_column: str          # Column name from file
    target_field: str         # Target field name (e.g., "part_number")
    confidence: float         # How confident we are in this mapping (0-1)
    sample_values: List[str]  # Sample values from this column


@dataclass
class ImportRow:
    """Single row of imported data."""
    row_number: int
    raw_data: Dict[str, str]           # Original column -> value
    mapped_data: Dict[str, str]        # target_field -> value
    validation_errors: List[str]       # Any validation errors
    pn_match: Optional[Dict[str, Any]] # Matched part number if found


@dataclass
class ImportPreview:
    """Preview of import file before processing."""
    file_type: ImportFileType
    total_rows: int
    column_mappings: List[ColumnMapping]
    preview_rows: List[ImportRow]
    unmapped_columns: List[str]
    validation_summary: Dict[str, int]


# =============================================================================
# Schema-Aware Column Matching
# =============================================================================


# Global schema matcher instance (lazy initialized)
_schema_matcher = None
_schema_provider = None


def _get_schema_matcher():
    """
    Get or create the SchemaColumnMatcher instance (lazy initialization).

    Uses lazy import to avoid cold start impact in AgentCore runtime.
    """
    global _schema_matcher
    if _schema_matcher is None:
        try:
            from tools.schema_column_matcher import get_column_matcher
            _schema_matcher = get_column_matcher()
            logger.info("[CSVParser] Schema-aware matching enabled")
        except Exception as e:
            logger.warning(f"[CSVParser] Schema matcher unavailable: {e}")
    return _schema_matcher


def _get_schema_provider():
    """
    Get or create the SchemaProvider instance (lazy initialization).
    """
    global _schema_provider
    if _schema_provider is None:
        try:
            from tools.schema_provider import get_schema_provider
            _schema_provider = get_schema_provider()
            logger.info("[CSVParser] Schema provider enabled")
        except Exception as e:
            logger.warning(f"[CSVParser] Schema provider unavailable: {e}")
    return _schema_provider


def get_expected_columns(target_table: str = "pending_entry_items") -> Dict[str, List[str]]:
    """
    Get expected columns from schema with aliases.

    Args:
        target_table: Target PostgreSQL table

    Returns:
        Dictionary of target_field → list of aliases
    """
    # Try schema-aware approach first
    provider = _get_schema_provider()
    if provider:
        try:
            schema = provider.get_table_schema(target_table)
            if schema:
                # Return column names from schema
                # Note: aliases come from SchemaColumnMatcher, not here
                columns = {}
                for col in schema.columns:
                    columns[col.name] = [col.name]
                return columns
        except Exception as e:
            logger.warning(f"[CSVParser] Failed to get schema columns: {e}")

    # Fallback to legacy patterns
    return EXPECTED_COLUMNS


def get_required_fields(target_table: str = "pending_entry_items") -> List[str]:
    """
    Get required fields from schema.

    Args:
        target_table: Target PostgreSQL table

    Returns:
        List of required field names
    """
    # Try schema-aware approach first
    provider = _get_schema_provider()
    if provider:
        try:
            return provider.get_required_columns(target_table)
        except Exception as e:
            logger.warning(f"[CSVParser] Failed to get required columns: {e}")

    # Fallback to legacy
    return REQUIRED_FIELDS


# Legacy expected columns (fallback when schema unavailable)
# Column names updated to match PostgreSQL schema
EXPECTED_COLUMNS: Dict[str, List[str]] = {
    "part_number": ["pn", "codigo", "material", "part_number", "partnumber", "cod_material", "codigo_material"],
    "description": ["desc", "descricao", "nome", "description", "nome_material", "desc_material"],
    "quantity": ["qty", "quantidade", "qtd", "quant", "quantity"],
    "serial_number": ["serial", "sn", "serie", "serial_number", "numero_serie"],
    "location_code": ["loc", "local", "deposito", "location", "warehouse", "armazem"],
    "project_code": ["projeto", "project", "proj", "id_projeto"],
    "supplier_name": ["cod_fornecedor", "supplier_code", "cod_forn", "fornecedor_cod", "fornecedor"],
    "unit_value": ["custo", "cost", "preco", "valor", "unit_cost", "custo_unitario"],
    "ncm": ["ncm", "ncm_code", "codigo_ncm"],
}

# Legacy required fields (fallback when schema unavailable)
REQUIRED_FIELDS = ["part_number", "quantity"]


# =============================================================================
# CSV Parser Functions
# =============================================================================

def detect_delimiter(sample: str) -> str:
    """
    Auto-detect CSV delimiter from sample.

    Args:
        sample: First few lines of the file

    Returns:
        Detected delimiter character
    """
    # Common delimiters to check
    delimiters = [",", ";", "\t", "|"]

    # Count occurrences in first line
    first_line = sample.split("\n")[0] if "\n" in sample else sample

    counts = {d: first_line.count(d) for d in delimiters}

    # Return delimiter with highest count (must be > 0)
    best = max(counts, key=counts.get)

    return best if counts[best] > 0 else ","


def normalize_column_name(name: str) -> str:
    """
    Normalize column name for matching.

    Args:
        name: Original column name

    Returns:
        Normalized lowercase name without special chars
    """
    # Remove accents (simple approach)
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "è": "e", "ê": "e",
        "í": "i", "ì": "i", "î": "i",
        "ó": "o", "ò": "o", "õ": "o", "ô": "o",
        "ú": "u", "ù": "u", "û": "u",
        "ç": "c",
    }

    result = name.lower().strip()
    for char, replacement in replacements.items():
        result = result.replace(char, replacement)

    # Remove special characters, keep alphanumeric and underscore
    result = "".join(c if c.isalnum() or c == "_" else "_" for c in result)

    # Remove multiple underscores
    while "__" in result:
        result = result.replace("__", "_")

    return result.strip("_")


def map_column_to_field(
    column_name: str,
    target_table: str = "pending_entry_items",
) -> Tuple[Optional[str], float]:
    """
    Map a column name to expected field using schema-aware matching.

    Algorithm (priority order):
    1. Use SchemaColumnMatcher (schema + aliases + fuzzy) if available
    2. Fallback to legacy EXPECTED_COLUMNS if schema unavailable

    Args:
        column_name: Original column name from file
        target_table: Target PostgreSQL table (default: pending_entry_items)

    Returns:
        Tuple of (target_field or None, confidence)
    """
    # Try schema-aware matching first
    matcher = _get_schema_matcher()
    if matcher:
        target, confidence = matcher.match_column(column_name, target_table)
        if target:
            logger.debug(
                f"[CSVParser] Schema match: {column_name} → {target} ({confidence:.2f})"
            )
            return target, confidence

    # Fallback to legacy pattern matching
    normalized = normalize_column_name(column_name)

    for field, aliases in EXPECTED_COLUMNS.items():
        # Exact match
        if normalized in [normalize_column_name(a) for a in aliases]:
            return field, 0.95

        # Partial match (contains)
        for alias in aliases:
            alias_norm = normalize_column_name(alias)
            if alias_norm in normalized or normalized in alias_norm:
                return field, 0.75

    return None, 0.0


def parse_csv_content(
    content: bytes,
    encoding: str = "utf-8",
    max_preview_rows: int = 10,
) -> ImportPreview:
    """
    Parse CSV content and return preview.

    Args:
        content: Raw file content as bytes
        encoding: File encoding (default utf-8)
        max_preview_rows: Number of rows to include in preview

    Returns:
        ImportPreview with mappings and sample data
    """
    # Decode content
    try:
        text = content.decode(encoding)
    except UnicodeDecodeError:
        # Try latin-1 as fallback (common for Brazilian files)
        text = content.decode("latin-1")

    # Detect delimiter
    delimiter = detect_delimiter(text)

    # Parse CSV
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    if not reader.fieldnames:
        raise ValueError("CSV file has no headers")

    # Map columns
    column_mappings: List[ColumnMapping] = []
    unmapped_columns: List[str] = []
    field_to_column: Dict[str, str] = {}

    for col in reader.fieldnames:
        target_field, confidence = map_column_to_field(col)

        if target_field and confidence >= 0.5:
            # Check if this field is already mapped
            if target_field not in field_to_column:
                column_mappings.append(ColumnMapping(
                    file_column=col,
                    target_field=target_field,
                    confidence=confidence,
                    sample_values=[],
                ))
                field_to_column[target_field] = col
            else:
                unmapped_columns.append(col)
        else:
            unmapped_columns.append(col)

    # Parse rows
    rows: List[ImportRow] = []
    all_rows = list(reader)

    for i, row_data in enumerate(all_rows[:max_preview_rows]):
        # Build mapped data
        mapped_data: Dict[str, str] = {}
        for mapping in column_mappings:
            value = row_data.get(mapping.file_column, "").strip()
            mapped_data[mapping.target_field] = value

            # Collect sample values
            if value and len(mapping.sample_values) < 3:
                mapping.sample_values.append(value)

        # Validate row
        errors: List[str] = []
        for required in REQUIRED_FIELDS:
            if not mapped_data.get(required):
                errors.append(f"Campo obrigatório faltando: {required}")

        # Validate quantity is numeric
        qty = mapped_data.get("quantity", "")
        if qty:
            try:
                float(qty.replace(",", "."))
            except ValueError:
                errors.append(f"Quantidade inválida: {qty}")

        rows.append(ImportRow(
            row_number=i + 2,  # +2 for header and 1-based index
            raw_data=dict(row_data),
            mapped_data=mapped_data,
            validation_errors=errors,
            pn_match=None,
        ))

    # Count validation issues
    validation_summary = {
        "valid_rows": sum(1 for r in rows if not r.validation_errors),
        "error_rows": sum(1 for r in rows if r.validation_errors),
        "total_rows": len(all_rows),
    }

    return ImportPreview(
        file_type=ImportFileType.CSV,
        total_rows=len(all_rows),
        column_mappings=column_mappings,
        preview_rows=rows,
        unmapped_columns=unmapped_columns,
        validation_summary=validation_summary,
    )


def parse_excel_content(
    content: bytes,
    sheet_name: Optional[str] = None,
    max_preview_rows: int = 10,
) -> ImportPreview:
    """
    Parse Excel content and return preview.

    NOTE: Requires openpyxl package. If not available, will raise ImportError.
    Consider adding openpyxl to requirements.txt if Excel support is needed.

    Args:
        content: Raw file content as bytes
        sheet_name: Specific sheet to parse (default: first sheet)
        max_preview_rows: Number of rows to include in preview

    Returns:
        ImportPreview with mappings and sample data
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl package required for Excel support. "
            "Add 'openpyxl' to requirements.txt"
        )

    # Load workbook from bytes
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Select sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Get headers from first row
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)

    if not headers:
        raise ValueError("Excel file has no headers")

    # Convert headers to strings
    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

    # Map columns (same logic as CSV)
    column_mappings: List[ColumnMapping] = []
    unmapped_columns: List[str] = []
    field_to_column: Dict[str, str] = {}

    for col in headers:
        target_field, confidence = map_column_to_field(col)

        if target_field and confidence >= 0.5:
            if target_field not in field_to_column:
                column_mappings.append(ColumnMapping(
                    file_column=col,
                    target_field=target_field,
                    confidence=confidence,
                    sample_values=[],
                ))
                field_to_column[target_field] = col
            else:
                unmapped_columns.append(col)
        else:
            unmapped_columns.append(col)

    # Parse rows
    rows: List[ImportRow] = []
    all_rows: List[tuple] = []

    for row_tuple in rows_iter:
        all_rows.append(row_tuple)

    for i, row_tuple in enumerate(all_rows[:max_preview_rows]):
        row_data = {headers[j]: str(v) if v else "" for j, v in enumerate(row_tuple)}

        # Build mapped data
        mapped_data: Dict[str, str] = {}
        for mapping in column_mappings:
            value = row_data.get(mapping.file_column, "").strip()
            mapped_data[mapping.target_field] = value

            if value and len(mapping.sample_values) < 3:
                mapping.sample_values.append(value)

        # Validate row
        errors: List[str] = []
        for required in REQUIRED_FIELDS:
            if not mapped_data.get(required):
                errors.append(f"Campo obrigatório faltando: {required}")

        qty = mapped_data.get("quantity", "")
        if qty:
            try:
                float(qty.replace(",", "."))
            except ValueError:
                errors.append(f"Quantidade inválida: {qty}")

        rows.append(ImportRow(
            row_number=i + 2,
            raw_data=row_data,
            mapped_data=mapped_data,
            validation_errors=errors,
            pn_match=None,
        ))

    wb.close()

    validation_summary = {
        "valid_rows": sum(1 for r in rows if not r.validation_errors),
        "error_rows": sum(1 for r in rows if r.validation_errors),
        "total_rows": len(all_rows),
    }

    return ImportPreview(
        file_type=ImportFileType.XLSX,
        total_rows=len(all_rows),
        column_mappings=column_mappings,
        preview_rows=rows,
        unmapped_columns=unmapped_columns,
        validation_summary=validation_summary,
    )


def parse_import_file(
    content: bytes,
    filename: str,
    max_preview_rows: int = 10,
) -> ImportPreview:
    """
    Parse import file (CSV or Excel) and return preview.

    Args:
        content: Raw file content as bytes
        filename: Original filename (used to detect type)
        max_preview_rows: Number of rows for preview

    Returns:
        ImportPreview with mappings and sample data
    """
    lower_name = filename.lower()

    if lower_name.endswith(".xlsx"):
        return parse_excel_content(content, max_preview_rows=max_preview_rows)
    elif lower_name.endswith(".xls"):
        raise ValueError("Formato .xls não suportado. Use .xlsx ou .csv")
    else:
        # Default to CSV
        return parse_csv_content(content, max_preview_rows=max_preview_rows)


# =============================================================================
# Export Preview to Dict (for JSON serialization)
# =============================================================================

def preview_to_dict(preview: ImportPreview) -> Dict[str, Any]:
    """
    Convert ImportPreview to dict for JSON serialization.

    Args:
        preview: ImportPreview object

    Returns:
        Dictionary representation
    """
    return {
        "file_type": preview.file_type.value,
        "total_rows": preview.total_rows,
        "column_mappings": [
            {
                "file_column": m.file_column,
                "target_field": m.target_field,
                "confidence": m.confidence,
                "sample_values": m.sample_values,
            }
            for m in preview.column_mappings
        ],
        "preview_rows": [
            {
                "row_number": r.row_number,
                "raw_data": r.raw_data,
                "mapped_data": r.mapped_data,
                "validation_errors": r.validation_errors,
                "pn_match": r.pn_match,
            }
            for r in preview.preview_rows
        ],
        "unmapped_columns": preview.unmapped_columns,
        "validation_summary": preview.validation_summary,
    }


# =============================================================================
# Bulk Data Extraction (for actual import)
# =============================================================================

def extract_all_rows(
    content: bytes,
    filename: str,
    column_mappings: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    """
    Extract all rows from file using specified column mappings.

    Args:
        content: Raw file content
        filename: Original filename
        column_mappings: List of {file_column, target_field} mappings

    Returns:
        List of mapped row dictionaries
    """
    lower_name = filename.lower()

    # Build mapping lookup
    mapping_lookup = {m["file_column"]: m["target_field"] for m in column_mappings}

    if lower_name.endswith(".xlsx"):
        return _extract_all_xlsx(content, mapping_lookup)
    else:
        return _extract_all_csv(content, mapping_lookup)


def _extract_all_csv(
    content: bytes,
    mapping_lookup: Dict[str, str],
) -> List[Dict[str, str]]:
    """Extract all rows from CSV."""
    # Decode
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    delimiter = detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    rows = []
    for row_data in reader:
        mapped = {}
        for file_col, target_field in mapping_lookup.items():
            value = row_data.get(file_col, "").strip()
            mapped[target_field] = value
        rows.append(mapped)

    return rows


def _extract_all_xlsx(
    content: bytes,
    mapping_lookup: Dict[str, str],
) -> List[Dict[str, str]]:
    """Extract all rows from Excel."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl required for Excel support")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)

    if not headers:
        wb.close()
        return []

    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

    rows = []
    for row_tuple in rows_iter:
        row_data = {headers[j]: str(v) if v else "" for j, v in enumerate(row_tuple)}

        mapped = {}
        for file_col, target_field in mapping_lookup.items():
            value = row_data.get(file_col, "").strip()
            mapped[target_field] = value
        rows.append(mapped)

    wb.close()
    return rows
