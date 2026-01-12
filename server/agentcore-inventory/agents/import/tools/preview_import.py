# =============================================================================
# Preview Import Tool
# =============================================================================
# Analyzes CSV/Excel files before import with column detection and PN matching.
# =============================================================================

import csv
import io
import logging
from typing import Dict, Any, List, Optional, Tuple
from difflib import SequenceMatcher


from shared.audit_emitter import AgentAuditEmitter
from shared.xray_tracer import trace_tool_call

logger = logging.getLogger(__name__)

AGENT_ID = "import"
audit = AgentAuditEmitter(agent_id=AGENT_ID)

# Column mapping patterns (source → target)
COLUMN_PATTERNS = {
    "part_number": [
        "codigo", "código", "part_number", "pn", "partnumber",
        "part", "item", "material", "sku", "cod", "code"
    ],
    "description": [
        "descricao", "descrição", "description", "desc", "nome",
        "name", "item_desc", "product_name", "produto"
    ],
    "quantity": [
        "quantidade", "qty", "qtd", "quant", "qtt", "qt",
        "amount", "count", "qtde"
    ],
    "serial_number": [
        "serial", "serial_number", "ns", "sn", "numero_serie",
        "série", "serie", "serialnumber"
    ],
    "location": [
        "localizacao", "localização", "location", "local", "loc",
        "storage", "armazem", "armazém", "warehouse", "deposito"
    ],
    "project_id": [
        "projeto", "project", "project_id", "proj", "cliente",
        "customer", "contrato", "contract"
    ],
    "unit_value": [
        "valor_unitario", "valor_unit", "unit_value", "preco",
        "price", "valor", "custo", "cost", "unit_price"
    ],
    "notes": [
        "observacao", "observação", "notes", "obs", "remarks",
        "comentario", "comment", "nota"
    ],
}


@trace_tool_call("sga_preview_import")
async def preview_import_tool(
    s3_key: str,
    filename: str,
    project_id: str,
    destination_location_id: str,
    session_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Preview import file and generate column mappings.

    Args:
        s3_key: S3 key of uploaded file
        filename: Original filename
        project_id: Target project ID
        destination_location_id: Target location ID
        session_id: Optional session ID for audit

    Returns:
        Preview with detected columns, mappings, and matched rows
    """
    audit.working(
        message=f"Analisando arquivo: {filename}",
        session_id=session_id,
    )

    try:
        # Download file from S3
        file_content, file_type = await _download_file(s3_key)

        if not file_content:
            return {
                "success": False,
                "error": f"Arquivo não encontrado: {s3_key}",
            }

        # Parse file based on type
        if file_type in ["xlsx", "xls"]:
            headers, rows = await _parse_excel(file_content, file_type)
        else:
            headers, rows = await _parse_csv(file_content)

        if not headers:
            return {
                "success": False,
                "error": "Não foi possível detectar colunas no arquivo",
            }

        # Detect column mappings
        column_mappings = await detect_columns_tool(
            headers=headers,
            session_id=session_id,
        )

        # Match rows to part numbers (first 100 for preview)
        preview_rows = rows[:100]
        matched_rows = await match_rows_to_pn(
            rows=preview_rows,
            column_mappings=column_mappings.get("mappings", []),
            session_id=session_id,
        )

        # Calculate statistics
        total_rows = len(rows)
        matched_count = sum(1 for r in matched_rows.get("rows", []) if r.get("matched"))
        match_rate = matched_count / len(preview_rows) if preview_rows else 0

        audit.completed(
            message=f"Preview gerado: {total_rows} linhas, {match_rate:.0%} match rate",
            session_id=session_id,
            details={
                "total_rows": total_rows,
                "matched": matched_count,
                "match_rate": match_rate,
            },
        )

        return {
            "success": True,
            "filename": filename,
            "file_type": file_type,
            "total_rows": total_rows,
            "headers": headers,
            "column_mappings": column_mappings.get("mappings", []),
            "unmapped_columns": column_mappings.get("unmapped", []),
            "preview_rows": matched_rows.get("rows", []),
            "match_rate": match_rate,
            "project_id": project_id,
            "destination_location_id": destination_location_id,
        }

    except Exception as e:
        logger.error(f"[preview_import] Error: {e}", exc_info=True)
        audit.error(
            message="Erro ao analisar arquivo",
            session_id=session_id,
            error=str(e),
        )
        return {
            "success": False,
            "error": str(e),
        }


@trace_tool_call("sga_detect_columns")
async def detect_columns_tool(
    headers: List[str],
    session_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Detect column mappings from headers.

    Args:
        headers: List of column headers from file
        session_id: Optional session ID for audit

    Returns:
        Detected mappings and unmapped columns
    """
    mappings = []
    unmapped = []

    for header in headers:
        if not header:
            continue

        normalized = _normalize_header(header)
        target_field, confidence = _match_column(normalized)

        if target_field and confidence >= 0.6:
            mappings.append({
                "source_column": header,
                "target_field": target_field,
                "confidence": confidence,
                "auto_detected": True,
            })
        else:
            unmapped.append(header)

    return {
        "success": True,
        "mappings": mappings,
        "unmapped": unmapped,
        "detected_count": len(mappings),
    }


@trace_tool_call("sga_match_rows_pn")
async def match_rows_to_pn(
    rows: List[Dict[str, Any]],
    column_mappings: Optional[List[Dict[str, Any]]] = None,
    session_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Match rows to part numbers in the database.

    Args:
        rows: Data rows from file
        column_mappings: Optional column mappings to apply
        session_id: Optional session ID for audit

    Returns:
        Rows with match results
    """
    if not rows:
        return {"success": True, "rows": []}

    # Get PN column from mappings
    pn_column = None
    desc_column = None

    if column_mappings:
        for mapping in column_mappings:
            if mapping.get("target_field") == "part_number":
                pn_column = mapping.get("source_column")
            elif mapping.get("target_field") == "description":
                desc_column = mapping.get("source_column")

    # Get all part numbers from database
    db_parts = await _get_part_numbers()

    matched_rows = []
    for row in rows:
        row_data = dict(row)

        # Try exact match by code
        pn_value = row.get(pn_column) if pn_column else None
        matched_pn = None
        match_confidence = 0
        match_method = None

        if pn_value:
            # Exact match
            for part in db_parts:
                if str(pn_value).strip().upper() == str(part.get("part_number", "")).strip().upper():
                    matched_pn = part["part_number"]
                    match_confidence = 1.0
                    match_method = "exact"
                    break

        # Fuzzy match by description if no exact match
        if not matched_pn and desc_column:
            desc_value = row.get(desc_column)
            if desc_value:
                best_match, best_score = _fuzzy_match_description(
                    desc_value,
                    db_parts,
                )
                if best_score >= 0.8:
                    matched_pn = best_match
                    match_confidence = best_score
                    match_method = "fuzzy"

        row_data["matched"] = matched_pn is not None
        row_data["matched_pn"] = matched_pn
        row_data["match_confidence"] = match_confidence
        row_data["match_method"] = match_method

        matched_rows.append(row_data)

    return {
        "success": True,
        "rows": matched_rows,
        "total": len(matched_rows),
        "matched_count": sum(1 for r in matched_rows if r.get("matched")),
    }


# =============================================================================
# Helper Functions
# =============================================================================

async def _download_file(s3_key: str) -> Tuple[Optional[bytes], str]:
    """Download file from S3 and detect type."""
    try:
        import boto3
        import os

        s3 = boto3.client("s3")
        bucket = os.environ.get("IMPORT_BUCKET", "faiston-one-imports-prod")

        response = s3.get_object(Bucket=bucket, Key=s3_key)
        content = response["Body"].read()

        # Detect file type from extension or magic bytes
        file_type = "csv"
        if s3_key.lower().endswith(".xlsx"):
            file_type = "xlsx"
        elif s3_key.lower().endswith(".xls"):
            file_type = "xls"
        elif content[:4] == b"PK\x03\x04":  # ZIP signature (XLSX)
            file_type = "xlsx"

        return content, file_type

    except Exception as e:
        logger.error(f"[preview_import] S3 download error: {e}")
        return None, "csv"


async def _parse_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parse CSV content."""
    try:
        # Try different encodings
        for encoding in ["utf-8", "latin-1", "cp1252"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode("utf-8", errors="replace")

        # Detect delimiter
        sample = text[:2000]
        delimiter = ","
        for delim in [";", "\t", "|"]:
            if sample.count(delim) > sample.count(delimiter):
                delimiter = delim

        # Parse CSV
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
        rows = list(reader)

        return headers, rows

    except Exception as e:
        logger.error(f"[preview_import] CSV parse error: {e}")
        return [], []


async def _parse_excel(content: bytes, file_type: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parse Excel content using openpyxl (lazy import)."""
    try:
        from openpyxl import load_workbook
        import io as io_module

        wb = load_workbook(filename=io_module.BytesIO(content), read_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        rows = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else "" for cell in row]
            else:
                if any(cell for cell in row):  # Skip empty rows
                    row_dict = {}
                    for j, cell in enumerate(row):
                        if j < len(headers) and headers[j]:
                            row_dict[headers[j]] = cell
                    rows.append(row_dict)

        return headers, rows

    except ImportError:
        logger.warning("[preview_import] openpyxl not available, falling back to CSV")
        return [], []
    except Exception as e:
        logger.error(f"[preview_import] Excel parse error: {e}")
        return [], []


def _normalize_header(header: str) -> str:
    """Normalize header for matching."""
    import re
    return re.sub(r"[^a-z0-9]", "", header.lower())


def _match_column(normalized: str) -> Tuple[Optional[str], float]:
    """Match normalized header to target field."""
    best_field = None
    best_score = 0

    for target_field, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            pattern_norm = _normalize_header(pattern)

            # Exact match
            if normalized == pattern_norm:
                return target_field, 1.0

            # Partial match
            score = SequenceMatcher(None, normalized, pattern_norm).ratio()
            if score > best_score:
                best_score = score
                best_field = target_field

    return best_field, best_score


def _fuzzy_match_description(
    description: str,
    parts: List[Dict[str, Any]],
) -> Tuple[Optional[str], float]:
    """Fuzzy match description to part numbers."""
    if not description or not parts:
        return None, 0

    best_match = None
    best_score = 0

    desc_lower = description.lower()

    for part in parts:
        part_desc = part.get("description", "")
        if not part_desc:
            continue

        score = SequenceMatcher(None, desc_lower, part_desc.lower()).ratio()
        if score > best_score:
            best_score = score
            best_match = part.get("part_number")

    return best_match, best_score


async def _get_part_numbers() -> List[Dict[str, Any]]:
    """Get all part numbers from database."""
    try:
        from tools.db_client import DBClient
        db = DBClient()
        return await db.list_part_numbers()
    except ImportError:
        logger.warning("[preview_import] DBClient not available")
        return []
    except Exception as e:
        logger.error(f"[preview_import] DB error: {e}")
        return []
